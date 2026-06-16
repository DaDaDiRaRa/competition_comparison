"""
P0-3 validation: run brief analysis pipeline on actual PDFs and verify
BRIEF_PROJECT_INFO classification + xlsx Sheet 1 "사업 개요" section.

Run from repo root:
    python tools/p0_3_validate.py
"""

import sys, os, json, io, tempfile, pathlib

# Add backend to path
BACKEND = pathlib.Path(__file__).resolve().parents[1] / "backend"
sys.path.insert(0, str(BACKEND))

# Minimal config to avoid app_settings.json dependency
os.environ.setdefault("ANTHROPIC_API_KEY", os.environ.get("ANTHROPIC_API_KEY", ""))

PDFS = {
    "jongno_main": r"M:\103_건원 NEW 프론티어 프로그램\2020년\D2_사내 디자인 공모전\02_종로구청 설계공모 지침서_KOR.pdf",
    "jongno_sub":  r"M:\103_건원 NEW 프론티어 프로그램\2020년\D2_사내 디자인 공모전\03_종로구청 시설별 세부지침서_KOR.pdf",
    "yeongdp":     r"D:\BACK DATA\competition_comparison\설계공모 수정지침서(영등포구 통합 신청사 국제설계공모)_25.07.18..pdf",
}

FACILITY_TYPE = "public"  # 공공시설 — general group

# ──────────────────────────────────────────────────────────────────────────────

def classify_brief(pdf_path: str, label: str):
    """Classify all pages in a brief PDF and return page_map."""
    import asyncio
    from services.page_classifier import classify_all_pages_brief
    from config import settings

    print(f"\n{'='*70}")
    print(f"[CLASSIFY] {label}")
    print(f"  PDF: {pdf_path}")

    api_key = settings.api_key
    if not api_key:
        print("  !! No API key — skipping classification (LLM call required)")
        return None

    # classify_all_pages_brief returns list[dict]: [{page: int, type: str, ...}, ...]
    results = asyncio.run(classify_all_pages_brief(pdf_path))
    page_map = results  # extract_pdf expects list[dict] as-is
    print(f"  Total pages classified: {len(page_map)}")

    bpi_pages = [item["page"] for item in page_map if item.get("primary_type") == "BRIEF_PROJECT_INFO"]
    print(f"  BRIEF_PROJECT_INFO pages: {bpi_pages if bpi_pages else '(none)'}")

    # Print full distribution
    from collections import Counter
    _NEW_TYPES = {
        "BRIEF_PROJECT_INFO", "BRIEF_DESIGN_MASSING", "BRIEF_DESIGN_FACADE",
        "BRIEF_DESIGN_SUSTAIN", "BRIEF_DESIGN_SPECIAL",
    }
    dist = Counter(item.get("primary_type") for item in page_map)
    for ptype, cnt in sorted(dist.items()):
        marker = " <-- NEW" if ptype in _NEW_TYPES else ""
        print(f"    {ptype}: {cnt}{marker}")

    # P2-3-1: BRIEF_DESIGN_* 분산 체크
    design_types = {
        k: v for k, v in dist.items()
        if k and k.startswith("BRIEF_DESIGN_")
    }
    total_design = sum(design_types.values())
    guide_cnt = design_types.get("BRIEF_DESIGN_GUIDE", 0)
    if total_design:
        pct = guide_cnt / total_design * 100
        print(f"\n  [P2-3] 설계 지침 분산: 총 {total_design}페이지")
        for dt, cnt in sorted(design_types.items()):
            print(f"    {dt}: {cnt}페이지")
        print(f"  BRIEF_DESIGN_GUIDE(폴백) 비율: {pct:.0f}%  {'<-- 높음(개선 필요)' if pct > 50 else 'OK'}")

    return page_map


def extract_brief(pdf_path: str, page_map: list, label: str):
    """Extract brief data and return brief_data dict."""
    import asyncio
    from services.data_extractor import extract_pdf, merge_extracted_data

    print(f"\n[EXTRACT] {label}")

    raw = asyncio.run(extract_pdf(pdf_path, page_map, is_brief=True))
    brief_data = merge_extracted_data(page_map, raw)

    # ── area_table (P1 new schema) — ALL brief_program pages ─────────────────
    _bp_all = brief_data.get("brief_program") or []
    if isinstance(_bp_all, dict):
        _bp_all = [_bp_all]
    area_table: list = []
    shared_areas: list = []
    for _bpp in _bp_all:
        if isinstance(_bpp, dict):
            area_table.extend(_bpp.get("area_table") or [])
            shared_areas.extend(_bpp.get("shared_areas") or [])

    if area_table:
        total_items = sum(len(g.get("items") or []) for g in area_table if isinstance(g, dict))
        total_subs  = sum(
            len(item.get("sub_items") or [])
            for g in area_table if isinstance(g, dict)
            for item in (g.get("items") or []) if isinstance(item, dict)
        )
        depth = 1 + (1 if total_items else 0) + (1 if total_subs else 0)
        print(f"  area_table: {len(area_table)} group(s)  items={total_items}  sub_items={total_subs}  depth={depth}단")
        for gi, grp in enumerate(area_table):
            if not isinstance(grp, dict):
                continue
            gname  = grp.get("group_name") or "(no name)"
            gtotal = grp.get("total_area_sqm")
            site   = grp.get("site_id") or "-"
            items  = grp.get("items") or []
            print(f"    [{gi}] '{gname}'  site={site}  total={gtotal}㎡  items={len(items)}")
            for ii, item in enumerate(items[:5]):  # max 5 items per group
                if not isinstance(item, dict):
                    continue
                subs = item.get("sub_items") or []
                print(f"         [{ii}] {item.get('name')}  {item.get('area_sqm')}㎡  sub_items={len(subs)}")
                for si, sub in enumerate(subs[:3]):  # max 3 subs
                    if isinstance(sub, dict):
                        print(f"              [{si}] {sub.get('name')}  {sub.get('area_sqm')}㎡")
            if len(items) > 5:
                print(f"         ... ({len(items)-5} more items)")
        if shared_areas:
            print(f"  shared_areas: {len(shared_areas)} item(s)")
            for sa in shared_areas[:3]:
                if isinstance(sa, dict):
                    print(f"    '{sa.get('name')}'  {sa.get('area_sqm')}㎡")
    else:
        print("  area_table: (not extracted — may be old rooms[] schema or empty BRIEF_PROGRAM)")

    # ── KI-2: BRIEF_PROJECT_INFO 수치 필드 체크 ──────────────────────────────
    def _first(d, k):
        v = d.get(k) or {}
        if isinstance(v, list): v = v[0] if v else {}
        return v if isinstance(v, dict) else {}

    bpi = _first(brief_data, "brief_project_info")
    if bpi:
        print(f"\n  [KI-2] BRIEF_PROJECT_INFO 수치 필드:")
        for field in ("competition_name", "organizer", "construction_cost_100m_won",
                      "construction_period_months"):
            print(f"    {field}: {bpi.get(field)}")
        for i, site in enumerate(bpi.get("sites") or []):
            if not isinstance(site, dict): continue
            print(f"    sites[{i}] {site.get('site_id')}: "
                  f"site_area={site.get('site_area_sqm')}  "
                  f"bcr={site.get('building_coverage_pct')}%  "
                  f"far={site.get('floor_area_ratio_pct')}%  "
                  f"height={site.get('max_height_m')}m")
    else:
        print("\n  [KI-2] BRIEF_PROJECT_INFO: (not extracted)")

    # ── KI-1: BRIEF_EVALUATION 배점 합계 체크 ────────────────────────────────
    be = _first(brief_data, "brief_evaluation")
    if be:
        total_pts = be.get("total_points")
        cats = be.get("evaluation_categories") or []
        computed = sum(
            c.get("points") or 0 for c in cats
            if isinstance(c, dict) and isinstance(c.get("points"), (int, float))
        )
        print(f"\n  [KI-1] BRIEF_EVALUATION:")
        print(f"    total_points(LLM): {total_pts}  computed_sum: {computed}  "
              f"{'OK' if 95 <= computed <= 105 else '<-- 이상 (중복 집계 의심)'}")
        for c in cats[:5]:
            if isinstance(c, dict):
                print(f"    {c.get('name')}: {c.get('points')}점")
        if len(cats) > 5:
            print(f"    ... ({len(cats)-5} more)")
    else:
        print("\n  [KI-1] BRIEF_EVALUATION: (not extracted)")

    # ── P2-3-3/4: BRIEF_DESIGN_* 추출 데이터 체크 ───────────────────────────
    print(f"\n  [P2-3] BRIEF_DESIGN_* 추출 결과:")
    for dtype, label_ko in [
        ("brief_design_massing", "배치·동선"),
        ("brief_design_facade",  "입면·재료"),
        ("brief_design_sustain", "친환경·인증"),
        ("brief_design_special", "특수·보안"),
        ("brief_design_guide",   "기타(폴백)"),
    ]:
        raw = brief_data.get(dtype)
        if not raw:
            print(f"    {dtype}: (없음)")
            continue
        pages = raw if isinstance(raw, list) else [raw]
        all_items = 0
        for page in pages:
            if not isinstance(page, dict): continue
            for v in page.values():
                if isinstance(v, list): all_items += len(v)
                elif v and not isinstance(v, (dict, list)): all_items += 1
        print(f"    {dtype} ({label_ko}): {len(pages)}페이지  총 추출항목~{all_items}개")
        # 인증 수치 (P2-3-3/4)
        if dtype == "brief_design_sustain":
            for page in pages:
                if not isinstance(page, dict): continue
                certs = page.get("required_certifications") or []
                rpct  = page.get("renewable_energy_min_pct")
                if certs:
                    for c in certs:
                        if isinstance(c, dict):
                            print(f"      인증: {c.get('name')}  등급: {c.get('required_grade')}")
                if rpct is not None:
                    print(f"      신재생에너지 최소비율: {rpct}%")

    return brief_data


def export_xlsx(brief_data: dict, label: str):
    """Generate xlsx and check Sheet 1 for 사업 개요 section."""
    from services.brief_checklist_exporter import to_xlsx

    print(f"\n[XLSX] {label}")

    # Use empty validation dict (no flags)
    validation = {"flags": [], "summary": {"high": 0, "medium": 0, "low": 0}, "checked_rules": []}
    raw_bytes = to_xlsx(brief_data, validation)

    # Save to temp file and inspect with openpyxl
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
    ws1 = wb.worksheets[0]

    print(f"  Sheet 1 name: {ws1.title}")
    print(f"  Rows: {ws1.max_row}  Cols: {ws1.max_column}")

    # Check key sections + col B values (기준면적A)
    found_area_prog = False
    found_area_limit = False
    row_data = []  # (col_A_text, col_B_value, col_B_indent)
    for r in ws1.iter_rows(min_row=1):
        cell0 = str(r[0].value or "").strip()
        cell1 = r[1].value if len(r) > 1 else None  # 기준면적(A)
        indent = getattr(r[0].alignment, "indent", 0) if r[0].alignment else 0
        if "실별 면적 프로그램" in cell0:
            found_area_prog = True
        if "전체 규모 한도" in cell0 or "부지별 건축개요" in cell0:
            found_area_limit = True
        if cell0:
            row_data.append((cell0, cell1, indent))

    print(f"  {'✓' if found_area_prog else '✗'} '실별 면적 프로그램' 섹션")
    print(f"  {'✓' if found_area_limit else '✗'} '전체 규모 한도' / '부지별 건축개요' 섹션")

    # Count rows with area_table data (non-empty col B in program section)
    in_prog = False
    filled_B = 0
    for lbl, colB, _ind in row_data:
        if "실별 면적 프로그램" in lbl:
            in_prog = True
        if in_prog and colB is not None:
            filled_B += 1
    print(f"  기준면적(A) 채워진 행: {filled_B}개")

    print(f"\n  Sheet 1 rows (col A  |  col B  |  indent) — first 35:")
    for lbl, colB, ind in row_data[:35]:
        indent_str = "  " * int(ind) if ind else ""
        print(f"    {indent_str}{lbl:<30}  |  {colB}")

    # ── P2-3-6: Sheet 3 설계 지침 섹션 체크 ─────────────────────────────────
    if len(wb.worksheets) >= 3:
        ws3 = wb.worksheets[2]
        _S3_SECTIONS = ["배치·동선", "입면·재료", "친환경·인증", "특수·보안", "기타 설계"]
        found_s3 = {s: False for s in _S3_SECTIONS}
        for r in ws3.iter_rows(min_row=1):
            cell0 = str(r[0].value or "").strip()
            for s in _S3_SECTIONS:
                if s in cell0:
                    found_s3[s] = True
        print(f"\n  Sheet 3 ({ws3.title}) 설계 지침 섹션:")
        for s, ok in found_s3.items():
            print(f"    {'✓' if ok else '✗'} '{s}' 섹션")

    # Save for manual inspection
    safe_label = label.replace("/", "_").replace("\\", "_")
    out_path = pathlib.Path(tempfile.gettempdir()) / f"p03_{safe_label}.xlsx"
    out_path.write_bytes(raw_bytes)
    print(f"\n  Saved to: {out_path}")
    return out_path


def export_md(brief_data: dict, label: str):
    """Generate markdown and check P2-3-5 section headers."""
    from services.brief_checklist_exporter import to_markdown
    print(f"\n[MD] {label}")
    validation = {"flags": [], "summary": {"high": 0, "medium": 0, "low": 0}, "checked_rules": []}
    md_text = to_markdown(brief_data, validation)

    _MD_SECTIONS = ["배치·동선 지침", "입면·재료 지침", "친환경·인증 요구사항", "특수·보안 지침", "기타 설계 지침"]
    print("  MD 설계 지침 섹션:")
    for s in _MD_SECTIONS:
        found = f"### {s}" in md_text
        print(f"    {'✓' if found else '✗'} '### {s}'")

    # Save MD
    safe_label = label.replace("/", "_").replace("\\", "_")
    out_path = pathlib.Path(tempfile.gettempdir()) / f"p03_{safe_label}.md"
    out_path.write_text(md_text, encoding="utf-8")
    print(f"  Saved to: {out_path}")


def check_page_map_from_saved_json(label: str, json_path: str):
    """If we have a saved _brief.json, check it directly without running LLM."""
    p = pathlib.Path(json_path)
    if not p.exists():
        return None
    print(f"\n[JSON] Loading saved brief JSON for {label}: {json_path}")
    data = json.loads(p.read_text(encoding="utf-8"))
    return data


def run_xlsx_only(label: str, brief_data: dict):
    """Skip classify/extract — just test xlsx generation from a dict."""
    return export_xlsx(brief_data, label)


# ──────────────────────────────────────────────────────────────────────────────

def main():
    from config import settings, BRIEF_PAGE_TYPES, BRIEF_PAGE_TYPES_META

    print("\n" + "="*70)
    print("P0-3 VALIDATION - BRIEF_PROJECT_INFO pipeline test")
    print("="*70)

    # Sanity check config
    _NEW_TYPES_CHECK = [
        "BRIEF_PROJECT_INFO", "BRIEF_DESIGN_MASSING", "BRIEF_DESIGN_FACADE",
        "BRIEF_DESIGN_SUSTAIN", "BRIEF_DESIGN_SPECIAL",
    ]
    print("\n[CONFIG] BRIEF_PAGE_TYPES 등록 확인:")
    for t in _NEW_TYPES_CHECK:
        ok = t in BRIEF_PAGE_TYPES
        meta = BRIEF_PAGE_TYPES_META.get(t, "(없음)")
        print(f"  {'✓' if ok else '✗'} {t}  meta={meta}")
    assert all(t in BRIEF_PAGE_TYPES for t in _NEW_TYPES_CHECK), "일부 타입 미등록!"

    # Check API key
    api_key = settings.api_key
    if not api_key:
        print("\n!! ANTHROPIC_API_KEY not set. Classification/extraction steps require LLM.")
        print("   Set env var and re-run, or provide saved _brief.json paths.")
        print("\n[XLSX-ONLY] Testing xlsx export with synthetic brief_data...")
        _test_xlsx_synthetic()
        return

    for label, pdf_path in PDFS.items():
        if not pathlib.Path(pdf_path).exists():
            print(f"\n[SKIP] {label} — file not found: {pdf_path}")
            continue

        page_map = classify_brief(pdf_path, label)
        if page_map is None:
            continue

        brief_data = extract_brief(pdf_path, page_map, label)
        export_xlsx(brief_data, label)
        export_md(brief_data, label)

    print("\n" + "="*70)
    print("P0-3 / P1-3 / P2-3 DONE")
    print("="*70)


def _test_xlsx_synthetic():
    """Test xlsx export with hand-crafted brief_data containing BRIEF_PROJECT_INFO."""
    # Single-site synthetic data
    brief_single = {
        "brief_project_info": {
            "competition_name": "종로구청 설계공모",
            "organizer": "종로구",
            "competition_type": "설계공모",
            "sites": [
                {
                    "site_id": "단일부지",
                    "address": "서울특별시 종로구 종로1·2·3·4가동",
                    "zoning": "제2종 일반주거지역",
                    "scope": "신축",
                    "facilities": ["구청 청사", "주민 편의시설"],
                    "site_area_sqm": 5200.0,
                    "floor_area_sqm": 18500.0,
                    "building_coverage_pct": 60.0,
                    "floor_area_ratio_pct": 350.0,
                    "max_height_m": 45.0,
                    "open_space_sqm": 800.0,
                    "open_space_notes": "공개공지 법정 기준 이상",
                }
            ],
            "construction_cost_100m_won": 320.0,
            "design_cost_100m_won": 12.5,
            "construction_period_months": 30,
            "budget_notes": ["VAT 별도", "공사비 산정 기준: 2024년 단가"],
            "special_conditions": ["면적 허용 오차 ±5%"],
        }
    }

    # Add brief_program with area_table (1-level simple)
    brief_single["brief_program"] = {
        "area_table": [
            {
                "group_name": "구청사",
                "site_id": "단일부지",
                "total_area_sqm": 12000.0,
                "items": [
                    {"name": "민원실", "area_sqm": 800.0, "notes": "1층 배치", "sub_items": []},
                    {"name": "사무실", "area_sqm": 4500.0, "notes": "", "sub_items": []},
                    {"name": "회의실", "area_sqm": 600.0, "notes": "2~3층", "sub_items": []},
                    {"name": "로비·홀", "area_sqm": 500.0, "notes": "", "sub_items": []},
                ],
            },
            {
                "group_name": "주민 편의시설",
                "site_id": "단일부지",
                "total_area_sqm": 2800.0,
                "items": [
                    {"name": "도서관", "area_sqm": 1200.0, "notes": "", "sub_items": []},
                    {"name": "체육관", "area_sqm": 1600.0, "notes": "", "sub_items": []},
                ],
            },
        ],
        "shared_areas": [
            {"name": "공용주차장", "area_sqm": 3200.0, "notes": "지하 1~2층"},
            {"name": "기계·전기실", "area_sqm": 500.0, "notes": ""},
        ],
    }

    # Multi-site synthetic data
    brief_multi = {
        "brief_project_info": {
            "competition_name": "영등포구 통합 신청사 국제설계공모",
            "organizer": "영등포구",
            "competition_type": "국제설계공모",
            "sites": [
                {
                    "site_id": "부지1",
                    "address": "서울특별시 영등포구 영등포동 A",
                    "zoning": "일반상업지역",
                    "scope": "신축",
                    "facilities": ["구청 본관"],
                    "site_area_sqm": 8400.0,
                    "floor_area_sqm": None,
                    "building_coverage_pct": 50.0,
                    "floor_area_ratio_pct": 800.0,
                    "max_height_m": 90.0,
                    "open_space_sqm": 1200.0,
                    "open_space_notes": "",
                },
                {
                    "site_id": "부지2",
                    "address": "서울특별시 영등포구 영등포동 B",
                    "zoning": "준주거지역",
                    "scope": "신축",
                    "facilities": ["주민 편의시설", "주차장"],
                    "site_area_sqm": 3100.0,
                    "floor_area_sqm": 9300.0,
                    "building_coverage_pct": 60.0,
                    "floor_area_ratio_pct": 300.0,
                    "max_height_m": 35.0,
                    "open_space_sqm": None,
                    "open_space_notes": "",
                },
            ],
            "construction_cost_100m_won": 980.0,
            "design_cost_100m_won": None,
            "construction_period_months": 48,
            "budget_notes": [],
            "special_conditions": ["각 부지 별도 발주", "면적 허용 오차 ±3%"],
        }
    }

    # Add brief_program with area_table (3-level complex) to multi-site
    brief_multi["brief_program"] = {
        "area_table": [
            {
                "group_name": "행정동 (부지1)",
                "site_id": "부지1",
                "total_area_sqm": 35000.0,
                "items": [
                    {
                        "name": "구청 본청",
                        "area_sqm": 20000.0,
                        "notes": "",
                        "sub_items": [
                            {"name": "민원창구", "area_sqm": 800.0, "notes": "1F"},
                            {"name": "행정사무실", "area_sqm": 12000.0, "notes": "2~8F"},
                            {"name": "대강당", "area_sqm": 1200.0, "notes": ""},
                        ],
                    },
                    {
                        "name": "구의회",
                        "area_sqm": 5000.0,
                        "notes": "별관 연결",
                        "sub_items": [
                            {"name": "의회 본회의장", "area_sqm": 800.0, "notes": ""},
                            {"name": "의원실", "area_sqm": 2000.0, "notes": ""},
                        ],
                    },
                ],
            },
            {
                "group_name": "생활SOC동 (부지2)",
                "site_id": "부지2",
                "total_area_sqm": 9300.0,
                "items": [
                    {"name": "주민센터", "area_sqm": 3000.0, "notes": "", "sub_items": []},
                    {"name": "도서관", "area_sqm": 2500.0, "notes": "", "sub_items": []},
                    {"name": "체육시설", "area_sqm": 3800.0, "notes": "", "sub_items": []},
                ],
            },
        ],
        "shared_areas": [
            {"name": "지하주차장 (공용)", "area_sqm": 12000.0, "notes": "B1~B3"},
        ],
    }

    print("\n  --- Synthetic single-site xlsx (1-level area_table) ---")
    out1 = export_xlsx(brief_single, "synthetic_single")

    print("\n  --- Synthetic multi-site xlsx (3-level area_table) ---")
    out2 = export_xlsx(brief_multi, "synthetic_multi")

    print(f"\n  Open these files to verify Sheet 1 '사업 개요' section:")
    print(f"    {out1}")
    print(f"    {out2}")


if __name__ == "__main__":
    main()
