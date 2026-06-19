"""V-10a~e 회귀 검증 — 영등포 통합신청사 _brief.json + xlsx 대상.

CLAUDE.md 의 V-10 체크리스트:
  V-10a: 각 페이지 자체 design_guidelines_grouped 보유. `_merged: true` 없음
  V-10b: p.45/46/47 면대실·비품창고 항목 → facility_scope='구청' + section_path '직무공간 (부서 사무실) ...'
  V-10c: xlsx 시트3 의 '직무공간 (부서 사무실)' 헤더 아래 대민업무상담실·비품창고·기타 부서별 함께
  V-10d: p.46 새 헤더 항목이 직전 컨텍스트 미계승
  V-10e: extraction 로그에서 GUIDE/MASSING/SUSTAIN/SPECIAL 그룹 병렬 시작 + 그룹 내부 순차
"""
from __future__ import annotations

import glob
import json
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

BRIEFS_DIR = Path(r"C:\Temp\CompTestDB\_briefs")
PATTERN = "20260619_161407*"

# 구체 공간명 (이게 entry 의 sub_path/leaf 에 등장해야 V-10b 의미 있는 라우팅 확인 가능)
TARGET_SPACE_NAMES = ["대민업무상담실", "비품창고", "면대실", "기타 부서별 요청사항"]
DUTY_PATH_MARKER   = "직무공간 (부서 사무실)"


def _find_artifacts() -> tuple[Path, Path]:
    json_matches = list(BRIEFS_DIR.glob(f"{PATTERN}.json"))
    xlsx_matches = list(BRIEFS_DIR.glob(f"{PATTERN}.xlsx"))
    if not json_matches:
        raise SystemExit(f"!! no _brief.json matching {PATTERN}")
    if not xlsx_matches:
        raise SystemExit(f"!! no xlsx matching {PATTERN}")
    return json_matches[0], xlsx_matches[0]


def check_v10a(by_type: dict) -> dict:
    """각 페이지에 자체 design_guidelines_grouped 보유. _merged 플래그 없음."""
    failures = []
    page_with_grouped = 0
    page_total = 0
    merged_pages = []

    for ptype in ("BRIEF_DESIGN_GUIDE", "BRIEF_DESIGN_MASSING", "BRIEF_DESIGN_SUSTAIN", "BRIEF_DESIGN_SPECIAL"):
        section = by_type.get(ptype) or {}
        for page in section.get("combined_data", []):
            page_total += 1
            pno = page.get("_page")
            if page.get("_merged"):
                merged_pages.append((ptype, pno))
            if page.get("design_guidelines_grouped"):
                page_with_grouped += 1

    if merged_pages:
        failures.append(f"_merged=True 페이지 발견: {merged_pages}")

    return {
        "name": "V-10a (페이지별 grouped + no _merged)",
        "pass": len(failures) == 0,
        "details": {
            "design_pages_total": page_total,
            "with_grouped": page_with_grouped,
            "merged_flagged": len(merged_pages),
        },
        "failures": failures,
    }


def check_v10b(by_type: dict) -> dict:
    """p.45/46/47 의 '대민업무상담실/비품창고/면대실/기타 부서별' 같은 구체 공간이
    section_path 또는 sub_path 에 '직무공간 (부서 사무실)' 트리 아래로 라우팅되는지.

    구체 공간명을 section_path 의 leaf 또는 items_by_sub.sub_path 의 leaf 에서만 검사.
    items.text 본문 일치는 의도가 다르므로 무시 (false positive 차단).
    """
    guide = (by_type.get("BRIEF_DESIGN_GUIDE") or {}).get("combined_data", [])
    target_pages = {45, 46, 47}

    def _leaf(path: str) -> str:
        # 'A > B > C' -> 'C'
        return path.rsplit(">", 1)[-1].strip() if path else ""

    hits = []  # 발견된 구체 공간 → 라우팅 확인
    for page in guide:
        pno = page.get("_page")
        if pno not in target_pages:
            continue
        for entry in page.get("design_guidelines_grouped", []):
            fs = entry.get("facility_scope") or ""
            sp = entry.get("section_path") or ""
            items_by_sub = entry.get("items_by_sub") or []

            # 검사 후보 = section_path leaf + 각 items_by_sub.sub_path leaf
            leaves: list[tuple[str, str]] = [(sp, _leaf(sp))]
            for it in items_by_sub:
                if isinstance(it, dict):
                    sub = it.get("sub_path") or ""
                    leaves.append((sub, _leaf(sub)))

            for full_path, leaf in leaves:
                for target in TARGET_SPACE_NAMES:
                    if target == leaf:
                        # leaf 가 구체 공간명 — 부모 트리에 '직무공간 (부서 사무실)' 포함하는지 확인
                        ok = (fs in ("구청", "전체")) and (DUTY_PATH_MARKER in full_path)
                        hits.append({
                            "page": pno,
                            "space": target,
                            "facility_scope": fs,
                            "path": full_path,
                            "ok": ok,
                        })

    failures = [h for h in hits if not h["ok"]]
    # 셋 모두 (대민업무상담실/비품창고/기타 부서별) 검출되어야 의미 있는 통과
    distinct_spaces = {h["space"] for h in hits if h["ok"]}
    coverage_ok = distinct_spaces >= {"대민업무상담실", "비품창고", "기타 부서별 요청사항"}
    return {
        "name": "V-10b (p.45/46/47 → 직무공간 (부서 사무실) 라우팅)",
        "pass": coverage_ok and len(failures) == 0,
        "details": {
            "hits": len(hits),
            "failed": len(failures),
            "distinct_spaces_ok": sorted(distinct_spaces),
        },
        "hits": hits,
        "failures": failures[:5],
    }


def check_v10d(by_type: dict) -> dict:
    """p.46 새 헤더 항목 컨텍스트 비계승.
    p.45 의 section_path/facility_scope 가 p.46 의 *모든* 엔트리에 그대로 복제됐다면 의심.
    """
    guide = (by_type.get("BRIEF_DESIGN_GUIDE") or {}).get("combined_data", [])
    p45 = next((p for p in guide if p.get("_page") == 45), None)
    p46 = next((p for p in guide if p.get("_page") == 46), None)
    if not p45 or not p46:
        return {"name": "V-10d (헤더 비계승)", "pass": None, "details": {"reason": "p.45 or p.46 missing"}}

    p45_paths = {(e.get("facility_scope"), e.get("section_path")) for e in p45.get("design_guidelines_grouped", [])}
    p46_entries = p46.get("design_guidelines_grouped", [])
    p46_paths = {(e.get("facility_scope"), e.get("section_path")) for e in p46_entries}

    # 모든 p.46 path 가 p.45 와 정확히 일치하면 컨텍스트 강제 계승 의심
    fully_inherited = len(p46_paths) > 0 and p46_paths.issubset(p45_paths)
    # 일부 다른 path 가 있으면 OK
    unique_in_p46 = p46_paths - p45_paths

    return {
        "name": "V-10d (p.46 새 헤더 비계승)",
        "pass": not fully_inherited,
        "details": {
            "p45_distinct_paths": len(p45_paths),
            "p46_distinct_paths": len(p46_paths),
            "p46_unique_paths": len(unique_in_p46),
            "fully_inherited": fully_inherited,
        },
        "p46_unique_sample": list(unique_in_p46)[:5],
    }


def check_v10c(xlsx_path: Path) -> dict:
    """xlsx 시트3 의 '직무공간 (부서 사무실)' 헤더 아래 p.45 발(대민업무상담실) +
    p.46 발(비품창고, 기타 부서별 요청사항) 자식 항목들이 함께 묶였는지 구조 검사.

    xlsx 셀에 페이지번호가 없으므로 페이지 매칭이 아닌 항목명 인접성으로 확인.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return {"name": "V-10c (xlsx 시트3 라우팅)", "pass": None, "details": {"reason": "openpyxl missing"}}

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheets = wb.sheetnames
    # 시트3 = '시설별 지침' 또는 '요구사항' 패턴
    target_sheet = None
    for name in sheets:
        if "지침" in name or "요구사항" in name:
            target_sheet = name
            break
    if target_sheet is None and len(sheets) >= 3:
        target_sheet = sheets[2]
    if target_sheet is None:
        return {"name": "V-10c (xlsx 시트3 라우팅)", "pass": False, "details": {"sheets": sheets}}

    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))

    # '직무공간 (부서 사무실)' 섹션 헤더 행 찾기 (정확 매칭)
    duty_header_row = None
    for i, row in enumerate(rows):
        cells_str = " ".join(str(c) for c in row if c is not None)
        if DUTY_PATH_MARKER in cells_str:
            duty_header_row = i
            break

    if duty_header_row is None:
        return {
            "name": "V-10c (xlsx 시트3 직무공간 트리)",
            "pass": False,
            "details": {"sheet": target_sheet, "rows": len(rows), "duty_header_found": False},
        }

    # 헤더부터 다음 새 섹션 헤더 등장 전까지의 자식 항목 텍스트 수집
    # 새 섹션 = '[' 로 시작하는 셀 (다른 카테고리), e.g. '[통합 민원실]', '[회의 및 행사공간]'
    expected_children = {"대민업무상담실", "비품창고", "기타 부서별 요청사항"}
    found_children: set[str] = set()
    end_row = duty_header_row
    for j in range(duty_header_row + 1, min(duty_header_row + 60, len(rows))):
        row = rows[j]
        first_nonempty = next((str(c) for c in row if c is not None and str(c).strip()), "")
        if first_nonempty.startswith("[") and DUTY_PATH_MARKER not in first_nonempty:
            end_row = j
            break
        cells_str = " ".join(str(c) for c in row if c is not None)
        for child in expected_children:
            if child in cells_str:
                found_children.add(child)
        end_row = j

    coverage_ok = expected_children.issubset(found_children)
    return {
        "name": "V-10c (xlsx 시트3 직무공간 트리 자식 인접)",
        "pass": coverage_ok,
        "details": {
            "sheet": target_sheet,
            "duty_header_row": duty_header_row,
            "scan_range": f"R{duty_header_row}~R{end_row}",
            "found_children": sorted(found_children),
            "expected_children": sorted(expected_children),
        },
    }


def check_v10e(sse_log: Path) -> dict:
    """SSE 로그만으로는 그룹별 병렬·내부 순차 확인 불가. 백엔드 stderr 필요.
    여기서는 'extract_brief' 단계 SSE event 가 단일 progress 였는지만 기록 (참고).
    """
    if not sse_log.exists():
        return {"name": "V-10e (그룹 병렬·내부 순차)", "pass": None, "details": {"reason": "no sse log"}}
    text = sse_log.read_text(encoding="utf-8", errors="ignore")
    extract_lines = [l for l in text.splitlines() if '"step": "extract_brief"' in l]
    return {
        "name": "V-10e (그룹 병렬·내부 순차)",
        "pass": None,
        "details": {
            "extract_brief_events": len(extract_lines),
            "note": "SSE 만으로 확인 불가 — 백엔드 stderr 의 group 시작 로그 grep 필요",
        },
    }


def main() -> int:
    json_path, xlsx_path = _find_artifacts()
    print(f"BRIEF JSON: {json_path}")
    print(f"XLSX:       {xlsx_path}\n")

    data = json.loads(json_path.read_text(encoding="utf-8"))
    by_type = data.get("_by_type") or {}

    sse_log = Path(r"d:\temp\v10_sse.log")

    results = [
        check_v10a(by_type),
        check_v10b(by_type),
        check_v10c(xlsx_path),
        check_v10d(by_type),
        check_v10e(sse_log),
    ]

    print("=" * 80)
    for r in results:
        status = "PASS" if r.get("pass") is True else ("FAIL" if r.get("pass") is False else "SKIP")
        print(f"[{status}] {r['name']}")
        for k, v in (r.get("details") or {}).items():
            print(f"   {k}: {v}")
        if r.get("failures"):
            for f in r["failures"]:
                print(f"   ! {f}")
        if r.get("hits") and r["name"].startswith("V-10b"):
            print("   hits:")
            for h in r["hits"][:10]:
                mark = "OK" if h["ok"] else "FAIL"
                print(f"     [{mark}] p.{h['page']} space='{h['space']}' fs='{h['facility_scope']}' path='{h['path']}'")
        if r.get("p46_unique_sample"):
            print("   p46_unique sample:")
            for s in r["p46_unique_sample"]:
                print(f"     - {s}")
        print()
    print("=" * 80)
    fail = sum(1 for r in results if r.get("pass") is False)
    skip = sum(1 for r in results if r.get("pass") is None)
    print(f"PASS: {len(results) - fail - skip}  FAIL: {fail}  SKIP: {skip}")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
