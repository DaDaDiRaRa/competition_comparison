"""
brief_checklist_exporter.py — 지침서 체크리스트 내보내기 (LLM 호출 없음, 렌더만)

공개 API:
  to_markdown(brief_data, validation) -> str   — Markdown 텍스트
  to_xlsx(brief_data, validation) -> bytes     — Excel 파일 (openpyxl)
  to_html(brief_data, validation) -> str       — 미니멀 스타일 HTML (보기/인쇄용)

인수:
  brief_data  : _brief.json dict
                (merge_extracted_data + _requirements + validation 머지 완료 상태)
  validation  : brief_data.get("validation", {})
                {"flags": [...], "summary": {high,medium,low}, "checked_rules": [...]}

섹션 (md/html 5섹션, 셋 다 _extract_sections 공유):
  1) 사업 개요 — brief_project_info (공모명·발주처·부지별 개요·사업 규모)
  2) 면적 프로그램 — area_rows/area_table + 면적·비율·층수 한도
  3) 심사기준 — evaluation_categories 배점표
  4) 요구사항·설계 지침 — requirements 축별 + design_guidelines_grouped
  5) 검증 경고 — validation.flags (severity 정렬 + xlsx 색 강조)
xlsx 는 1) 을 Sheet 1 서브섹션으로 접어 4시트(+ area_rows 시 5.면적표상세).

데이터 경로:
  새 BRIEF taxonomy  : brief_program / brief_regulations / brief_evaluation / brief_design_guide
  구 AREA_TABLE 경로 : area_table (room_program, zone_summary)
  _requirements      : extract_brief_requirements() 결과
  _quantitative      : merge_extracted_data() 자동 집계
"""
from __future__ import annotations

import html
import io
import re
from datetime import datetime
from typing import Any

from services.utils import _first, _as_list, normalize_design_guidelines_grouped  # 공유 dict 헬퍼


# ── 내부 헬퍼 ─────────────────────────────────────────────────────────────────

def _collect(data: dict, key: str, *list_keys: str) -> dict[str, list]:
    """다중 페이지 타입에서 list_key별 값을 모든 페이지에서 집계.

    Returns {list_key: [aggregated items]} — 한 페이지뿐이어도 동일하게 동작.
    """
    raw = data.get(key) or []
    if isinstance(raw, dict):
        raw = [raw]
    result: dict[str, list] = {lk: [] for lk in list_keys}
    for page in raw:
        if isinstance(page, dict):
            for lk in list_keys:
                result[lk].extend(page.get(lk) or [])
    return result


def _str_item(item: Any) -> str:
    """리스트 항목을 문자열로 변환. dict면 공통 텍스트 키 탐색."""
    if isinstance(item, str):
        return item
    if isinstance(item, dict):
        for k in ("description", "text", "name", "requirement", "guideline", "item"):
            if item.get(k):
                return str(item[k])
        return str(item)
    return str(item) if item is not None else ""


def _cell_safe(val: Any) -> Any:
    """openpyxl이 허용하는 스칼라로 변환. dict/list → 문자열 (숫자 dict는 합계)."""
    if val is None or isinstance(val, (bool, int, float, str)):
        return val
    if isinstance(val, dict):
        values = list(val.values())
        if values and all(isinstance(v, (int, float)) for v in values):
            return sum(values)
        return ", ".join(f"{k}: {v}" for k, v in val.items())
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    return str(val)


def _fmt_num(val: Any, unit: str = "") -> str:
    """숫자를 천 단위 구분 포맷으로 변환. None이면 빈 문자열."""
    if val is None:
        return ""
    if isinstance(val, dict):
        values = list(val.values())
        if values and all(isinstance(v, (int, float)) for v in values):
            val = sum(values)
        else:
            return ", ".join(f"{k}: {v}" for k, v in val.items())
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    try:
        fval = float(val)
        formatted = f"{int(fval):,}" if fval == int(fval) else f"{fval:,.1f}"
    except (TypeError, ValueError):
        return str(val)
    return f"{formatted}{unit}" if unit else formatted


_SEVERITY_LABEL = {"high": "높음", "medium": "보통", "low": "낮음"}
_SEVERITY_ORDER = {"high": 0, "medium": 1, "low": 2}

# 본문 면적표 절 번호 — "3.", "3-3." 등으로 시작하면 양식 모드 해제
_BODY_HDR_RE = re.compile(r"^\s*\d+\s*[.\-]")


def _form_area_pages(brief_data: dict) -> set:
    """'[서식 N] …면적표' 제출 양식으로 BRIEF_PROGRAM 에 오분류된 페이지 번호 집합.

    분류기가 제출 양식(예: '[서식 13] 건축 세부 면적표')을 BRIEF_PROGRAM 으로 잡으면
    본문 공간구성 면적표와 같은 실이 두 번 노출됨 (영등포 통합신청사 사례).
    page_header_text 에 '서식' 이 나오면 양식 시작, 본문 절 번호(3., 3-3.)나
    '공간구성' 이 나오면 본문 복귀. 연속(계속) 페이지는 직전 상태 유지.
    page_map 없으면 빈 집합 (안전 — 제외 없음).
    """
    pm = {p.get("page"): (p.get("page_header_text") or "")
          for p in (brief_data.get("page_map") or []) if isinstance(p, dict)}
    if not pm:
        return set()
    prog = brief_data.get("brief_program") or []
    if isinstance(prog, dict):
        prog = [prog]
    pages = sorted(p.get("_page") for p in prog
                   if isinstance(p, dict) and p.get("_page") is not None)
    excluded: set = set()
    in_form = False
    for pno in pages:
        h = pm.get(pno, "")
        if "서식" in h:
            in_form = True
        elif _BODY_HDR_RE.match(h) or "공간구성" in h:
            in_form = False
        if in_form:
            excluded.add(pno)
    return excluded


# ── 섹션 데이터 추출 ──────────────────────────────────────────────────────────

def _extract_sections(brief_data: dict) -> dict:
    """brief_data에서 4개 섹션용 정규화 데이터를 추출."""
    bp    = _first(brief_data, "brief_program")
    br    = _first(brief_data, "brief_regulations")
    at    = _first(brief_data, "area_table")
    # brief_evaluation: 배점이 가장 많은 페이지를 우선 사용.
    # 스태킹 폴백 시 개별 추출된 여러 페이지 중 실제 배점표가 있는 페이지를 선택.
    _be_pages = brief_data.get("brief_evaluation") or []
    if not isinstance(_be_pages, list):
        _be_pages = [_be_pages] if isinstance(_be_pages, dict) else []
    _be_pages = [p for p in _be_pages if isinstance(p, dict) and not p.get("_merged")]
    def _eval_pts(p: dict) -> int:
        cats = (p or {}).get("evaluation_categories") or []
        return sum(1 for c in cats if isinstance(c.get("points"), (int, float)))
    be = max(_be_pages, key=_eval_pts, default={})
    dg    = _first(brief_data, "brief_design_guide")   # 기타/폴백
    dm    = _first(brief_data, "brief_design_massing")
    dfa   = _first(brief_data, "brief_design_facade")
    ds    = _first(brief_data, "brief_design_sustain")
    bpi   = _first(brief_data, "brief_project_info")
    # BRIEF_DESIGN_* 다중 페이지 집계 — _first는 스칼라 전용, 리스트는 _collect 사용
    _dsp = _collect(brief_data, "brief_design_special",
                    "security_requirements", "accessibility_requirements",
                    "safety_requirements", "special_technical_requirements")
    _dm  = _collect(brief_data, "brief_design_massing",
                    "open_space_requirements", "parking_requirements",
                    "pedestrian_requirements", "connection_requirements",
                    "massing_guidelines")
    _dfa = _collect(brief_data, "brief_design_facade",
                    "primary_materials", "prohibited_materials",
                    "color_requirements", "facade_guidelines", "landscape_requirements")
    _ds  = _collect(brief_data, "brief_design_sustain",
                    "required_certifications", "energy_guidelines", "sustainability_requirements")
    quant = brief_data.get("_quantitative") or {}
    reqs  = brief_data.get("_requirements") or {}

    # ── Section 1: 면적·프로그램 ─────────────────────────────────────────────
    # sites 배열: 새 스키마(복수 부지 지원). 구 데이터는 빈 배열.
    sites = _as_list(bp, "sites")
    s0 = sites[0] if sites and isinstance(sites[0], dict) else {}

    total_fa = (
        bp.get("total_required_floor_area_sqm")
        or at.get("total_required_area_sqm")
        or quant.get("total_floor_area_sqm")
    )
    # 단일값: 구 top-level → sites[0] → area_table → quant 순
    site_area = (
        bp.get("site_area_sqm")
        or s0.get("site_area_sqm")
        or at.get("site_area_sqm")
        or quant.get("site_area_sqm")
    )
    # BRIEF_PROJECT_INFO sites 폴백 — 복수 부지면 "부지1: 60% / 부지2: 50%" 형식
    _bpi_sites = _as_list(bpi, "sites")

    def _bpi_pct(field: str):
        vals = [
            (st.get("site_id") or f"부지{i+1}", st.get(field))
            for i, st in enumerate(_bpi_sites)
            if isinstance(st, dict) and st.get(field) is not None
        ]
        if not vals:
            return None
        if len(vals) == 1:
            return vals[0][1]
        return " / ".join(f"{sid}: {v}%" for sid, v in vals)

    bcr = (
        bp.get("building_coverage_limit_pct")
        or s0.get("building_coverage_limit_pct")
        or br.get("building_coverage_ratio_limit_pct")
        or at.get("building_coverage_limit_pct")
        or quant.get("building_coverage_ratio_pct")
        or _bpi_pct("building_coverage_pct")
    )
    far = (
        bp.get("floor_area_ratio_limit_pct")
        or s0.get("floor_area_ratio_limit_pct")
        or br.get("floor_area_ratio_limit_pct")
        or at.get("floor_area_ratio_limit_pct")
        or quant.get("floor_area_ratio_pct")
        or _bpi_pct("floor_area_ratio_pct")
    )
    height       = br.get("height_limit_m") or s0.get("max_height_m") or dg.get("height_limit_m")
    floors_above = bp.get("max_floors_above") or quant.get("floors_above")
    floors_below = bp.get("max_floors_below") or quant.get("floors_below")
    parking      = (
        bp.get("required_parking")
        or at.get("parking_required")
        or quant.get("parking_count")
    )

    # 계층 면적표 — ALL brief_program 페이지의 area_table / area_rows 합산.
    # 단, '[서식 N] 건축 세부 면적표' 같은 제출 양식이 BRIEF_PROGRAM 으로 오분류된
    # 페이지는 본문 면적표와 같은 실을 중복 노출하므로 제외 (CLAUDE.md Known Issue).
    _bp_all = brief_data.get("brief_program") or []
    if isinstance(_bp_all, dict):
        _bp_all = [_bp_all]
    _excl = _form_area_pages(brief_data)
    area_table: list = []
    area_rows: list = []   # 신규 flat 방식
    shared_areas: list = []
    for _bpp in _bp_all:
        if isinstance(_bpp, dict) and _bpp.get("_page") not in _excl:
            area_table.extend(_bpp.get("area_table") or [])
            area_rows.extend(_bpp.get("area_rows") or [])
            shared_areas.extend(_bpp.get("shared_areas") or [])

    # 개략공사비 내역서 등 공사비 그룹 제거 (면적표와 무관)
    _COST_KW = {"공사비", "내역서", "공종", "원가", "견적"}
    area_table = [g for g in area_table
                  if not any(kw in (g.get("group_name") or "") for kw in _COST_KW)]

    # 구 경로 폴백: rooms / zones (area_table 없는 기존 데이터 호환)
    rooms = _as_list(bp, "rooms") or _as_list(at, "room_program")
    zones = _as_list(bp, "zones") or _as_list(at, "zone_summary")

    # 사업비·기간 — project-wide (병합 셀로 두 부지 공통)
    construction_cost = bp.get("estimated_construction_cost") or ""
    design_fee        = bp.get("estimated_design_fee") or ""
    design_period     = bp.get("design_period") or ""

    # ── Section 2: 심사기준 ───────────────────────────────────────────────────
    categories   = _as_list(be, "evaluation_categories")
    legacy_crit  = _as_list(reqs, "evaluation_criteria")
    eval_rows    = categories if categories else [
        {"name": c.get("item", ""), "points": c.get("points"), "description": ""}
        for c in legacy_crit
    ]
    total_points = be.get("total_points")
    eval_method  = be.get("evaluation_method") or ""
    jury         = be.get("jury_composition") or ""
    disqualify   = _as_list(be, "disqualification_criteria")

    # ── Section 3: 요구사항 ───────────────────────────────────────────────────
    requirements   = _as_list(reqs, "requirements")
    special_reqs   = _as_list(reqs, "special_requirements")
    # 기타/폴백 BRIEF_DESIGN_GUIDE (구 데이터 하위호환)
    design_reqs    = _as_list(dg, "design_requirements")
    setbacks       = _as_list(dg, "setback_requirements")
    materials      = _as_list(dg, "materials_required")
    sustainability = _as_list(dg, "sustainability_requirements")
    prohibited     = _as_list(dg, "prohibited_items")
    concept        = dg.get("concept_direction") or ""
    special_guide  = _as_list(dg, "special_guidelines")

    return {
        "area": {
            "total_fa": total_fa, "site_area": site_area,
            "bcr": bcr, "far": far, "height": height,
            "floors_above": floors_above, "floors_below": floors_below,
            "parking": parking,
            "area_rows": area_rows,                                    # 신규 flat 방식
            "area_table": area_table, "shared_areas": shared_areas,  # 새 계층 구조
            "rooms": rooms, "zones": zones,                           # 구 경로 폴백
            "sites": sites,  # 복수 부지 raw 배열 (단일 부지면 len==1 or [])
            "construction_cost": construction_cost,
            "design_fee": design_fee,
            "design_period": design_period,
        },
        "eval": {
            "rows": eval_rows, "total_points": total_points,
            "eval_method": eval_method, "jury": jury, "disqualify": disqualify,
            "points_sum_warning": bool(be.get("points_sum_warning")),
        },
        "reqs": {
            "requirements": requirements, "special_reqs": special_reqs,
            # 새 typed 설계 지침
            "massing": {
                "setback_m":    dm.get("building_setback_m"),
                "open_space":   _dm["open_space_requirements"],
                "parking":      _dm["parking_requirements"],
                "pedestrian":   _dm["pedestrian_requirements"],
                "connection":   _dm["connection_requirements"],
                "height_strategy": dm.get("height_strategy") or "",
                "guidelines":   _dm["massing_guidelines"],
            },
            "facade": {
                "primary_materials":   _dfa["primary_materials"],
                "prohibited_materials": _dfa["prohibited_materials"],
                "color":               _dfa["color_requirements"],
                "facade_guidelines":   _dfa["facade_guidelines"],
                "landscape":           _dfa["landscape_requirements"],
            },
            "sustain": {
                "certifications":    _ds["required_certifications"],
                "renewable_pct":     ds.get("renewable_energy_min_pct"),
                "energy_guidelines": _ds["energy_guidelines"],
                "sustainability_reqs": _ds["sustainability_requirements"],
            },
            "special": {
                "security":      _dsp["security_requirements"],
                "accessibility": _dsp["accessibility_requirements"],
                "safety":        _dsp["safety_requirements"],
                "special_tech":  _dsp["special_technical_requirements"],
            },
            # 기타/폴백 (구 BRIEF_DESIGN_GUIDE 하위호환)
            "design_reqs": design_reqs, "setbacks": setbacks,
            "materials": materials, "sustainability": sustainability,
            "prohibited": prohibited, "concept": concept,
            "special_guide": special_guide,
        },
        "project_info": {
            "sites": _as_list(bpi, "sites"),
            "competition_name": bpi.get("competition_name") or "",
            "organizer": bpi.get("organizer") or "",
            "competition_type": bpi.get("competition_type") or "",
            "construction_cost_100m_won": bpi.get("construction_cost_100m_won"),
            "design_cost_100m_won": bpi.get("design_cost_100m_won"),
            "construction_period_months": bpi.get("construction_period_months"),
            "budget_notes": _as_list(bpi, "budget_notes"),
            "special_conditions": _as_list(bpi, "special_conditions"),
            "unit_program": _as_list(bpi, "unit_program"),
        },
        # 계층 보존 설계지침 (merge_extracted_data 가 집계한 단일 리스트)
        # facility_scope/space_scope/category/section_path 기준으로 정렬·그룹화는 렌더링 단계에서 수행.
        # Lazy fallback: items_by_sub 가 없는 옛 형식이면 여기서 한 번 정규화 (호환).
        "guidelines_grouped": _ensure_normalized_grouped(brief_data.get("design_guidelines_grouped")),
    }


def _ensure_normalized_grouped(grouped: list | None) -> list:
    """이미 정규화된 데이터(items_by_sub 있음) 면 그대로, 아니면 정규화 적용."""
    if not grouped:
        return []
    if isinstance(grouped, list) and grouped and isinstance(grouped[0], dict) \
            and "items_by_sub" in grouped[0]:
        return grouped
    return normalize_design_guidelines_grouped(grouped)


# ── Markdown (구조화 데이터 덤프) ─────────────────────────────────────────────
# 테이블 포맷 없음 — key: value + 중첩 리스트로 데이터 밀도 최대화.
# null 필드도 "(없음)" 명시 → downstream 프로그램이 "누락"과 "미존재"를 구별 가능.

def _v(val: Any, unit: str = "") -> str:
    """값을 문자열로. None/빈값이면 '(없음)'. 이미 unit으로 끝나면 중복 추가 안 함."""
    if val is None or val == "" or val == []:
        return "(없음)"
    if isinstance(val, list):
        return ", ".join(str(x) for x in val) or "(없음)"
    s = str(val)
    if not unit or s == "(없음)" or s.endswith(unit):
        return s
    return s + unit


def to_markdown(brief_data: dict, validation: dict) -> str:
    """지침서 추출 데이터를 구조화 텍스트 덤프로 반환.

    가독성보다 데이터 밀도·분류 우선 — LLM/프로그램 연동용.
    """
    s  = _extract_sections(brief_data)
    a, e, r = s["area"], s["eval"], s["reqs"]
    pi = s["project_info"]
    flags = sorted(
        validation.get("flags") or [],
        key=lambda f: _SEVERITY_ORDER.get(f.get("severity", "low"), 2),
    )
    summary = validation.get("summary") or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    L: list[str] = [f"# 지침서 추출 데이터", f"생성: {now}", ""]

    # ══════════════════════════════════════════════════════════════════════════
    # 1. 사업 개요
    # ══════════════════════════════════════════════════════════════════════════
    L.append("## 1. 사업 개요")
    L.append(f"공모명: {_v(pi['competition_name'])}")
    L.append(f"발주처: {_v(pi['organizer'])}")
    L.append(f"공모유형: {_v(pi['competition_type'])}")
    L.append(f"예정공사비: {_v(pi['construction_cost_100m_won'], ' 억원')}")
    L.append(f"예정설계비: {_v(pi['design_cost_100m_won'], ' 억원')}")
    L.append(f"공사기간: {_v(pi['construction_period_months'], ' 개월')}")

    if pi["budget_notes"]:
        L.append("예산산정기준:")
        for n in pi["budget_notes"]:
            L.append(f"- {_str_item(n)}")

    if pi["special_conditions"]:
        L.append("특기사항:")
        for c in pi["special_conditions"]:
            L.append(f"- {_str_item(c)}")

    if pi["unit_program"]:
        L.append("단위세대·시설별 분배:")
        for u in pi["unit_program"]:
            if not isinstance(u, dict):
                L.append(f"- {_str_item(u)}")
                continue
            block      = u.get("block") or ""
            tenure     = u.get("tenure") or ""
            type_label = u.get("type_label") or ""
            area_text  = u.get("area_text") or ""
            ratio_text = u.get("ratio_text") or ""
            note       = u.get("note") or ""
            head = block + (f"({tenure})" if tenure else "")
            if type_label:
                head = (head + " · " if head else "") + type_label
            parts: list[str] = []
            if area_text:
                parts.append(area_text)
            if ratio_text:
                parts.append(ratio_text)
            if note:
                parts.append(f"비고: {note}")
            body = " / ".join(parts) if parts else "(내용 없음)"
            L.append(f"- {head}: {body}" if head else f"- {body}")

    # ── BRIEF_PROJECT_INFO 부지별 ─────────────────────────────────────────────
    for i, st in enumerate([x for x in pi["sites"] if isinstance(x, dict)]):
        sid = st.get("site_id") or f"부지{i+1}"
        L.append(f"\n### 부지개요: {sid}")
        L.append(f"위치: {_v(st.get('address'))}")
        L.append(f"용도지역지구: {_v(st.get('zoning'))}")
        L.append(f"공모범위: {_v(st.get('scope'))}")
        fac = st.get("facilities") or []
        L.append(f"도입시설: {', '.join(fac) if fac else '(없음)'}")
        L.append(f"대지면적: {_v(st.get('site_area_sqm'), ' ㎡')}")
        L.append(f"연면적: {_v(st.get('floor_area_sqm'), ' ㎡')}")
        L.append(f"건폐율: {_v(st.get('building_coverage_pct'), '%')}")
        L.append(f"용적률: {_v(st.get('floor_area_ratio_pct'), '%')}")
        L.append(f"최고높이: {_v(st.get('max_height_m'), ' m')}")
        L.append(f"공개공지: {_v(st.get('open_space_sqm'), ' ㎡')}")
        L.append(f"공개공지조건: {_v(st.get('open_space_notes'))}")

    # ── BRIEF_PROGRAM 부지별 ──────────────────────────────────────────────────
    bp_sites = [x for x in a["sites"] if isinstance(x, dict)]
    for i, st in enumerate(bp_sites):
        sid = st.get("site_id") or f"부지{i+1}"
        L.append(f"\n### 건축개요: {sid}")
        L.append(f"위치: {_v(st.get('address'))}")
        zoning = st.get("zoning") or []
        L.append(f"지역지구: {', '.join(zoning) if isinstance(zoning, list) else _v(zoning)}")
        L.append(f"건축구분: {_v(st.get('construction_type'))}")
        L.append(f"건축용도: {_v(st.get('building_use'))}")
        fac = st.get("facilities") or []
        L.append(f"도입시설: {', '.join(fac) if fac else '(없음)'}")
        L.append(f"대지면적: {_v(st.get('site_area_sqm'), ' ㎡')}")
        L.append(f"연면적: {_v(st.get('floor_area_sqm'), ' ㎡')}")
        L.append(f"건폐율한도: {_v(st.get('building_coverage_limit_pct'), '%')}")
        L.append(f"용적률한도: {_v(st.get('floor_area_ratio_limit_pct'), '%')}")
        L.append(f"최고높이: {_v(st.get('max_height_m'), ' m')}")
        L.append(f"공개공지: {_v(st.get('public_open_space_sqm'), ' ㎡')}")
        L.append(f"공개공지조건: {_v(st.get('public_open_space_notes'))}")

    L.append("\n### 전체 규모 한도")
    L.append(f"요구연면적: {_v(a['total_fa'], ' ㎡')}")
    L.append(f"대지면적: {_v(a['site_area'], ' ㎡')}")
    L.append(f"건폐율한도: {_v(a['bcr'], '%')}")
    L.append(f"용적률한도: {_v(a['far'], '%')}")
    L.append(f"높이한도: {_v(a['height'], ' m')}")
    L.append(f"지상층수: {_v(a['floors_above'], ' 층')}")
    L.append(f"지하층수: {_v(a['floors_below'], ' 층')}")
    L.append(f"주차대수: {_v(a['parking'], ' 대')}")
    L.append(f"예정공사비: {_v(a['construction_cost'])}")
    L.append(f"예정설계비: {_v(a['design_fee'])}")
    L.append(f"설계기간: {_v(a['design_period'])}")

    # ══════════════════════════════════════════════════════════════════════════
    # 2. 면적 프로그램
    # ══════════════════════════════════════════════════════════════════════════
    L.append("\n## 2. 면적 프로그램")

    _LEVEL_INDENT_MD = {"site_total": 0, "facility": 1, "bureau": 2, "division": 3, "space": 4}
    _LEVEL_LABEL_MD  = {0: "부지", 1: "시설", 2: "영역", 3: "과", 4: "세부"}

    if a["area_rows"]:
        for ar in a["area_rows"]:
            if not isinstance(ar, dict):
                continue
            rt      = ar.get("row_type") or "space"
            level   = _LEVEL_INDENT_MD.get(rt, 4)
            lbl     = _LEVEL_LABEL_MD.get(level, "")
            indent  = "  " * level
            name    = ar.get("name") or ""
            area_raw = ar.get("area") if ar.get("area") is not None else ar.get("subtotal_area")
            area_str = _v(area_raw, " ㎡") if area_raw is not None else ""
            notes   = ar.get("notes") or ""
            dept    = ar.get("dept") or ""
            sub_flag = " [소계]" if ar.get("is_subtotal") else ""
            extra   = ""
            if notes:
                extra += f"  # {notes}"
            if dept:
                extra += f"  소관:{dept}"
            L.append(f"{indent}- [{lbl}] {name}: {area_str}{sub_flag}{extra}")

    elif a["area_table"]:
        for grp in a["area_table"]:
            if not isinstance(grp, dict):
                continue
            gname  = grp.get("group_name") or "(무제)"
            gtotal = _v(grp.get("total_area_sqm"), " ㎡")
            gsite  = grp.get("site_id") or "-"
            L.append(f"\n### {gname}")
            L.append(f"부지: {gsite}")
            L.append(f"합계면적: {gtotal}")
            items = grp.get("items") or []
            if items:
                L.append("항목:")
                for it in items:
                    if not isinstance(it, dict):
                        continue
                    iname = it.get("name") or ""
                    iarea = _v(it.get("area_sqm"), " ㎡")
                    inote = it.get("notes") or ""
                    L.append(f"- {iname}: {iarea}" + (f"  # {inote}" if inote else ""))
                    for sub in (it.get("sub_items") or []):
                        if not isinstance(sub, dict):
                            continue
                        sname = sub.get("name") or ""
                        sarea = _v(sub.get("area_sqm"), " ㎡")
                        snote = sub.get("notes") or ""
                        L.append(f"  - {sname}: {sarea}" + (f"  # {snote}" if snote else ""))

        if a["shared_areas"]:
            L.append("\n### 공용·공동 면적")
            for sa in a["shared_areas"]:
                if not isinstance(sa, dict):
                    continue
                L.append(f"- {sa.get('name') or ''}: {_v(sa.get('area_sqm'), ' ㎡')}"
                         + (f"  # {sa.get('notes')}" if sa.get("notes") else ""))

    elif a["rooms"]:
        L.append("\n### 실별 면적 프로그램")
        for rm in a["rooms"]:
            if not isinstance(rm, dict):
                continue
            area = rm.get("required_area_sqm") or rm.get("area_sqm")
            L.append(f"- {rm.get('name') or ''}: {_v(area, ' ㎡')}"
                     f"  수량:{rm.get('required_count') or rm.get('count') or 1}"
                     + (f"  위치:{rm.get('floor')}" if rm.get("floor") else "")
                     + (f"  # {rm.get('notes')}" if rm.get("notes") else ""))

    if not a["area_table"] and a["zones"]:
        L.append("\n### 존 구성")
        for z in a["zones"]:
            L.append(f"- {z.get('name') or z.get('zone') or ''}: {_v(z.get('area_sqm'), ' ㎡')}")

    # ══════════════════════════════════════════════════════════════════════════
    # 3. 심사기준
    # ══════════════════════════════════════════════════════════════════════════
    L.append("\n## 3. 심사기준")
    L.append(f"총배점: {_v(e['total_points'], ' 점')}")
    L.append(f"평가방법: {_v(e['eval_method'])}")
    L.append(f"심사단구성: {_v(e['jury'])}")

    # 동일 이름 항목 병합: 표에 같은 구분이 여러 행으로 나뉜 경우(예: 10+5점) 하나로 합산
    _merged_rows: list[dict] = []
    _seen_names: dict[str, int] = {}  # name → index in _merged_rows
    for ev in e["rows"]:
        if not isinstance(ev, dict):
            continue
        name = ev.get("name") or "(항목명 없음)"
        if name in _seen_names:
            existing = _merged_rows[_seen_names[name]]
            # 배점 합산
            ep, np_ = existing.get("points"), ev.get("points")
            if isinstance(ep, (int, float)) and isinstance(np_, (int, float)):
                existing["points"] = ep + np_
            elif np_ is not None:
                existing["points"] = np_
            # sub_items 합산
            existing.setdefault("sub_items", [])
            existing["sub_items"].extend(ev.get("sub_items") or [])
            # shared_with 합산
            sw = existing.get("shared_with") or []
            sw.extend(sw2 for sw2 in (ev.get("shared_with") or []) if sw2 not in sw)
            existing["shared_with"] = sw
        else:
            _seen_names[name] = len(_merged_rows)
            _merged_rows.append(dict(ev))

    running = 0.0
    for ev in _merged_rows:
        name      = ev.get("name") or "(항목명 없음)"
        pts       = ev.get("points")
        shared    = ev.get("shared_with") or []
        sub_items = ev.get("sub_items") or []
        desc      = ev.get("description") or ""
        L.append(f"\n### {name}")
        L.append(f"배점: {_v(pts, ' 점')}")
        L.append(f"공유배점: {', '.join(shared) if shared else '(없음)'}")
        if desc:
            L.append(f"설명: {desc}")
        if sub_items:
            L.append("세부기준:")
            for sub in sub_items:
                L.append(f"- {_str_item(sub)}")
        if isinstance(pts, (int, float)):
            running += pts

    L.append(f"\n배점합계: {running if running else '(없음)'}")

    if e["disqualify"]:
        L.append("\n### 실격 요건")
        for d in e["disqualify"]:
            L.append(f"- {_str_item(d)}")

    # ══════════════════════════════════════════════════════════════════════════
    # 4. 요구사항·설계 지침
    # ══════════════════════════════════════════════════════════════════════════
    L.append("\n## 4. 요구사항·설계 지침")

    if r["requirements"]:
        L.append("\n### 평가축별 요구사항")
        for req in r["requirements"]:
            if not isinstance(req, dict):
                continue
            L.append(f"- 축: {req.get('axis') or ''}  "
                     f"배점비중: {_v(req.get('weight_pct'), '%')}  "
                     f"설명: {req.get('description') or ''}")

    if r["concept"]:
        L.append(f"\n설계방향: {r['concept']}")

    # ── 기타 설계 지침 (구 데이터 폴백) ─────────────────────────────────────
    # 계층 보존 설계지침 (시설별 / 공통 — design_guidelines_grouped 출처)
    grouped_all = s.get("guidelines_grouped") or []
    if grouped_all:
        # facility_scope 별 정렬 (멀티 시설 우선, 그 다음 "전체")
        facility_specific = [g for g in grouped_all
                             if (g.get("facility_scope") or "전체") != "전체"]
        common_grouped   = [g for g in grouped_all
                             if (g.get("facility_scope") or "전체") == "전체"]

        def _md_grouped_block(rows: list[dict], heading: str) -> None:
            if not rows:
                return
            L.append(f"\n### {heading}")
            # facility_scope 순서 유지 그룹화
            order: list[str] = []
            by_fac: dict[str, list[dict]] = {}
            for g in rows:
                fs = (g.get("facility_scope") or "전체").strip() or "전체"
                if fs not in by_fac:
                    order.append(fs)
                    by_fac[fs] = []
                by_fac[fs].append(g)
            for fs in order:
                if fs != "전체":
                    L.append(f"\n#### [{fs}]")
                for g in by_fac[fs]:
                    space    = (g.get("space_scope") or "전체").strip() or "전체"
                    cat      = (g.get("category") or "기타").strip() or "기타"
                    sec_path = (g.get("section_path") or "").strip()
                    # 정규화 후: items_by_sub 가 권위 데이터. 폴백으로 items 도 처리.
                    subs = g.get("items_by_sub")
                    if not subs:
                        items = g.get("items") or []
                        if not items:
                            continue
                        subs = [{"sub_path": "", "items": items}]
                    # 그룹 헤더 (한 번만)
                    head_parts: list[str] = []
                    if space != "전체":
                        head_parts.append(f"[{space}]")
                    head_parts.append(cat)
                    if sec_path:
                        head_parts.append(f"— {sec_path}")
                    L.append(f"\n**{' '.join(head_parts)}**")
                    # sub_path 별로 inline sub-header + items
                    for sub in subs:
                        sub_path = (sub.get("sub_path") or "").strip()
                        sub_items = sub.get("items") or []
                        if not sub_items:
                            continue
                        if sub_path:
                            # inline sub-header: "- 비품창고" 형태로 형제 레벨
                            L.append(f"- {sub_path}")
                            for it in sub_items:
                                if not isinstance(it, dict):
                                    continue
                                label = (it.get("label") or "").strip()
                                text  = (it.get("text") or "").strip()
                                if not text:
                                    continue
                                # label="-" 면 MD bullet 과 중복 → 생략
                                if not label or label == "-":
                                    L.append(f"  - {text}")
                                else:
                                    L.append(f"  - {label} {text}")
                        else:
                            for it in sub_items:
                                if not isinstance(it, dict):
                                    continue
                                label = (it.get("label") or "").strip()
                                text  = (it.get("text") or "").strip()
                                if not text:
                                    continue
                                if not label or label == "-":
                                    L.append(f"- {text}")
                                else:
                                    L.append(f"- {label} {text}")

        _md_grouped_block(facility_specific, "시설별 지침")
        _md_grouped_block(common_grouped,   "설계 지침 및 요구사항")

    if not grouped_all:
        _misc = [
            ("특수요구사항", r["special_reqs"]),
            ("기타설계지침", r["design_reqs"]),
            ("후퇴선요건",   r["setbacks"]),
            ("재료요건",     r["materials"]),
            ("친환경요건",   r["sustainability"]),
            ("금지사항",     r["prohibited"]),
            ("특별지침",     r["special_guide"]),
        ]
        has_misc = any(v for _, v in _misc)
        if has_misc:
            L.append("\n### 기타 설계 지침")
            for lbl, lst in _misc:
                if lst:
                    L.append(f"{lbl}:")
                    for x in lst:
                        L.append(f"- {_str_item(x)}")

    # ══════════════════════════════════════════════════════════════════════════
    # 5. 검증 경고
    # ══════════════════════════════════════════════════════════════════════════
    high_n   = summary.get("high", 0)
    medium_n = summary.get("medium", 0)
    low_n    = summary.get("low", 0)
    L.append(f"\n## 5. 검증 경고")
    L.append(f"높음: {high_n}건 / 보통: {medium_n}건 / 낮음: {low_n}건")

    if flags:
        for flag in flags:
            sev = _SEVERITY_LABEL.get(flag.get("severity", ""), flag.get("severity", ""))
            L.append(f"[{sev}] {flag.get('type') or ''}: {flag.get('message') or ''}"
                     + (f"  | 위치: {flag.get('location')}" if flag.get("location") else ""))
    else:
        L.append("경고없음")

    return "\n".join(L)


# ── HTML (미니멀 스타일, 보기용) ──────────────────────────────────────────────
# to_markdown 과 동일한 _extract_sections 데이터를 렌더 — 콘텐츠 정합성 보장.
# LLM 호출 없음. 자체 완결 HTML (외부 CSS/JS 의존 없음). 미니멀 화이트 + 건원 RED 포인트.

_HTML_CSS = """
:root{
  --ink:#1a1a1a; --text:#3a3a3a; --muted:#9a9a9a; --line:#ececec;
  --soft:#f7f7f8; --accent:#e60012;
  --high:#e60012; --med:#c47b00; --low:#9a9a9a;
}
*{box-sizing:border-box}
html{-webkit-text-size-adjust:100%}
body{margin:0;background:#fff;color:var(--text);
  font-family:'Apple SD Gothic Neo','Malgun Gothic',system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;
  font-size:14px;line-height:1.65;-webkit-font-smoothing:antialiased}
.wrap{max-width:900px;margin:0 auto;padding:52px 30px 110px}
header.doc{margin-bottom:8px;padding-bottom:22px;border-bottom:2px solid var(--ink)}
header.doc .eyebrow{font-size:12px;letter-spacing:.14em;color:var(--accent);font-weight:700;text-transform:uppercase}
header.doc h1{margin:8px 0 0;font-size:25px;font-weight:700;color:var(--ink);letter-spacing:-.02em;line-height:1.3}
header.doc .meta{margin-top:12px;color:var(--muted);font-size:12.5px;display:flex;flex-wrap:wrap;gap:6px 18px}
section.sec{margin:44px 0 0}
section.sec>h2{display:flex;align-items:center;gap:11px;margin:0 0 16px;
  font-size:18px;font-weight:700;color:var(--ink);letter-spacing:-.01em}
section.sec>h2 .n{display:inline-flex;align-items:center;justify-content:center;
  min-width:25px;height:25px;padding:0 7px;border-radius:7px;
  background:var(--accent);color:#fff;font-size:13px;font-weight:700;flex:0 0 auto}
h3.sub{margin:26px 0 9px;font-size:15px;font-weight:700;color:var(--ink);
  padding-bottom:7px;border-bottom:1px solid var(--line)}
h4.subsub{margin:18px 0 7px;font-size:13.5px;font-weight:600;color:var(--accent)}
.note{margin:14px 0;color:var(--text)}
dl.kv{display:grid;grid-template-columns:140px 1fr;gap:1px 18px;margin:6px 0}
dl.kv dt{color:var(--muted);font-size:13px;padding:4px 0}
dl.kv dd{margin:0;color:var(--ink);padding:4px 0;word-break:break-word}
dl.kv dd.empty{color:#c4c4c4}
table.t{width:100%;border-collapse:collapse;margin:10px 0;font-size:13px}
table.t th,table.t td{text-align:left;padding:9px 11px;border-bottom:1px solid var(--line);vertical-align:top}
table.t thead th{background:var(--soft);color:var(--muted);font-weight:600;font-size:11.5px;letter-spacing:.02em;
  position:sticky;top:0}
table.t td.num,table.t th.num{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}
table.t tr.sumrow td{font-weight:700;color:var(--ink);background:var(--soft);border-top:2px solid #ddd}
ul.list{margin:6px 0;padding-left:19px}
ul.list li{margin:3px 0}
.muted{color:var(--muted)}
.empty{color:#c4c4c4}
.tree{margin:8px 0;border-top:1px solid var(--line)}
.tree .row{display:flex;align-items:baseline;gap:10px;padding:5px 0;border-bottom:1px solid #f4f4f4}
.tree .row .lv{flex:0 0 42px;font-size:10.5px;color:var(--muted);text-align:right;padding-top:2px;
  letter-spacing:.04em}
.tree .row .nm{flex:1 1 auto;color:var(--ink)}
.tree .row .nm .meta{color:var(--muted);font-size:12px;margin-left:6px}
.tree .row .ar{flex:0 0 auto;font-variant-numeric:tabular-nums;color:var(--text);white-space:nowrap}
.tree .row.subtotal .nm,.tree .row.subtotal .ar{font-weight:700}
.gblock{margin:12px 0 4px}
.gblock .ghead{font-weight:700;color:var(--ink);margin-bottom:3px}
.gblock .gsub{color:var(--muted);font-weight:600;margin:8px 0 2px}
.chips{display:flex;flex-wrap:wrap;gap:6px;margin:6px 0}
.chip{background:var(--soft);border:1px solid var(--line);border-radius:14px;padding:3px 11px;font-size:12.5px;color:var(--ink)}
.warn-sum{display:flex;gap:8px;flex-wrap:wrap;margin:6px 0 14px}
.warn-sum .pill{display:inline-flex;align-items:center;gap:6px;border:1px solid var(--line);border-radius:20px;
  padding:5px 13px;font-size:12.5px;color:var(--text)}
.warn-sum .pill b{font-weight:700;color:var(--ink)}
.warn-sum .pill .dot{width:8px;height:8px;border-radius:50%;background:var(--low)}
.warn-sum .pill.high .dot{background:var(--high)} .warn-sum .pill.med .dot{background:var(--med)}
.warn-list .w{border-left:3px solid var(--low);background:var(--soft);
  padding:11px 14px;margin:9px 0;border-radius:0 7px 7px 0}
.warn-list .w.high{border-color:var(--high)} .warn-list .w.medium{border-color:var(--med)}
.warn-list .w .wt{font-weight:700;color:var(--ink);font-size:12.5px}
.warn-list .w .wt .sev{font-size:11px;color:var(--muted);font-weight:600;margin-left:7px}
.warn-list .w .wm{margin-top:3px;color:var(--text)}
.warn-list .w .wl{margin-top:4px;font-size:11.5px;color:var(--muted)}
.ok{color:#2a8a3e;font-weight:600}
html{scroll-behavior:smooth}
section.sec{scroll-margin-top:62px}
nav.top{position:sticky;top:0;z-index:50;background:rgba(255,255,255,.92);
  backdrop-filter:saturate(160%) blur(8px);-webkit-backdrop-filter:saturate(160%) blur(8px);
  border-bottom:1px solid var(--line)}
nav.top .inner{max-width:900px;margin:0 auto;padding:10px 30px;display:flex;align-items:center;gap:14px}
nav.top .ttl{font-weight:700;color:var(--ink);font-size:13px;white-space:nowrap;overflow:hidden;
  text-overflow:ellipsis;max-width:36%}
nav.top .links{display:flex;gap:2px;margin-left:auto;flex-wrap:nowrap;overflow-x:auto}
nav.top .links a{font-size:12.5px;color:var(--muted);text-decoration:none;padding:5px 11px;border-radius:7px;white-space:nowrap}
nav.top .links a:hover{background:var(--soft);color:var(--ink)}
.summary{display:grid;grid-template-columns:repeat(auto-fit,minmax(132px,1fr));gap:10px;margin:26px 0 6px}
.summary .card{border:1px solid var(--line);border-radius:11px;padding:13px 15px;background:#fff}
.summary .card .k{font-size:11.5px;color:var(--muted);margin-bottom:6px}
.summary .card .v{font-size:18px;font-weight:700;color:var(--ink);letter-spacing:-.01em;font-variant-numeric:tabular-nums;line-height:1.15}
.summary .card .v .u{font-size:12px;font-weight:500;color:var(--muted);margin-left:3px}
.summary .card.warn .v{color:var(--accent)}
.summary .card .sub{font-size:11px;color:var(--muted);margin-top:4px}
.totalbar{display:flex;justify-content:space-between;align-items:baseline;gap:12px;
  padding:9px 4px;margin:14px 0 2px;font-weight:700;color:var(--ink);border-bottom:2px solid #e0e0e0}
.totalbar .ar{font-variant-numeric:tabular-nums;white-space:nowrap}
details.fac{border:1px solid var(--line);border-radius:10px;margin:8px 0;overflow:hidden}
details.fac>summary{list-style:none;cursor:pointer;padding:11px 14px;display:flex;align-items:center;gap:11px;
  background:var(--soft);font-weight:600;color:var(--ink);user-select:none}
details.fac>summary::-webkit-details-marker{display:none}
details.fac>summary .caret{transition:transform .15s ease;color:var(--muted);flex:0 0 auto;font-size:10px}
details.fac[open]>summary .caret{transform:rotate(90deg)}
details.fac>summary .fname{flex:1 1 auto}
details.fac>summary .fmeta{color:var(--muted);font-weight:500;font-size:12.5px;font-variant-numeric:tabular-nums;white-space:nowrap}
details.fac>summary:hover{background:#f0f0f1}
details.fac .body{padding:2px 14px 8px}
details.fac .body .tree{border-top:none}
footer.doc{margin-top:64px;padding-top:18px;border-top:1px solid var(--line);color:#c0c0c0;font-size:11.5px;text-align:center}
@media print{.wrap{padding:0}body{font-size:12px}section.sec{break-inside:avoid}nav.top{display:none}details.fac>.body{display:block!important}details.fac>summary .caret{display:none}}
@media(max-width:560px){dl.kv{grid-template-columns:1fr}dl.kv dt{padding-bottom:0}.wrap{padding:32px 18px 64px}
  nav.top .ttl{display:none}nav.top .inner{padding:8px 18px}}
"""


def _esc(v: Any) -> str:
    """HTML 이스케이프. None/빈값은 빈 문자열."""
    if v is None:
        return ""
    return html.escape(str(v), quote=True)


def to_html(brief_data: dict, validation: dict) -> str:
    """지침서 추출 데이터를 미니멀 스타일 HTML 문서로 반환 (보기·인쇄용)."""
    s  = _extract_sections(brief_data)
    a, e, r = s["area"], s["eval"], s["reqs"]
    pi = s["project_info"]
    flags = sorted(
        validation.get("flags") or [],
        key=lambda f: _SEVERITY_ORDER.get(f.get("severity", "low"), 2),
    )
    summary = validation.get("summary") or {}
    bm = brief_data.get("_brief_meta") or {}
    # 제목: 문서 추출 공모명 우선 (사용자 라벨 brief_name 은 인코딩 깨짐 케이스가 있어 폴백).
    title = pi["competition_name"] or bm.get("brief_name") or "지침서 분석"
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    P: list[str] = []  # body HTML 조각

    def kv(pairs: list[tuple[str, Any]], units: dict | None = None) -> None:
        """key-value 정의 리스트. value 는 _v 로 포맷 ('(없음)' 처리)."""
        units = units or {}
        cells = []
        for lbl, val in pairs:
            txt = _v(val, units.get(lbl, ""))
            empty = txt == "(없음)"
            cells.append(f"<dt>{_esc(lbl)}</dt>"
                         f'<dd class="{"empty" if empty else ""}">{_esc(txt)}</dd>')
        P.append('<dl class="kv">' + "".join(cells) + "</dl>")

    def ul(items: list, fmt=_str_item) -> None:
        lis = [f"<li>{_esc(fmt(it))}</li>" for it in items if (fmt(it) or "").strip()]
        if lis:
            P.append('<ul class="list">' + "".join(lis) + "</ul>")

    # ══ Header ════════════════════════════════════════════════════════════════
    src = (bm.get("source_format") or "").upper()
    meta_bits = [f"<span>생성 {now}</span>"]
    if src:
        meta_bits.append(f"<span>원본 {_esc(src)}</span>")
    if brief_data.get("total_pages"):
        unit = "블록" if src == "DOCX" else "페이지"
        meta_bits.append(f"<span>{brief_data['total_pages']} {unit}</span>")
    P.append(
        '<header class="doc"><div class="eyebrow">설계공모 지침서 분석</div>'
        f"<h1>{_esc(title)}</h1>"
        f'<div class="meta">{"".join(meta_bits)}</div></header>'
    )

    # ══ 핵심 수치 요약 카드 ══════════════════════════════════════════════════════
    def _num(v) -> str:
        if isinstance(v, bool):
            return str(v)
        if isinstance(v, (int, float)):
            return f"{int(v):,}" if float(v).is_integer() else f"{v:,.1f}"
        return str(v)

    high_n = summary.get("high", 0)
    medium_n = summary.get("medium", 0)
    low_n = summary.get("low", 0)
    cards = []
    for k, val, unit in [
        ("연면적", a["total_fa"], "㎡"), ("대지면적", a["site_area"], "㎡"),
        ("용적률", a["far"], "%"), ("건폐율", a["bcr"], "%"),
        ("지상층수", a["floors_above"], "층"), ("지하층수", a["floors_below"], "층"),
        ("주차", a["parking"], "대"), ("총배점", e["total_points"], "점"),
    ]:
        if val is None or val == "":
            continue
        cards.append(f'<div class="card"><div class="k">{_esc(k)}</div>'
                     f'<div class="v">{_esc(_num(val))}<span class="u">{_esc(unit)}</span></div></div>')
    warn_total = high_n + medium_n + low_n
    cards.append(
        f'<div class="card{" warn" if high_n else ""}"><div class="k">검증 경고</div>'
        f'<div class="v">{warn_total}<span class="u">건</span></div>'
        f'<div class="sub">높음 {high_n} · 보통 {medium_n} · 낮음 {low_n}</div></div>'
    )
    P.append('<div class="summary">' + "".join(cards) + "</div>")

    nav_items: list[tuple[str, str]] = []  # (anchor, 짧은 라벨)

    def sec(num: int, heading: str, short: str) -> None:
        nav_items.append((f"s{num}", short))
        P.append(f'<section id="s{num}" class="sec"><h2>'
                 f'<span class="n">{num}</span>{_esc(heading)}</h2>')

    def endsec() -> None:
        P.append("</section>")

    # ══ 1. 사업 개요 ════════════════════════════════════════════════════════════
    sec(1, "사업 개요", "개요")
    kv([
        ("공모명", pi["competition_name"]),
        ("발주처", pi["organizer"]),
        ("공모유형", pi["competition_type"]),
        ("예정공사비", pi["construction_cost_100m_won"]),
        ("예정설계비", pi["design_cost_100m_won"]),
        ("공사기간", pi["construction_period_months"]),
    ], units={"예정공사비": " 억원", "예정설계비": " 억원", "공사기간": " 개월"})

    if pi["budget_notes"]:
        P.append('<h3 class="sub">예산 산정 기준</h3>')
        ul(pi["budget_notes"])
    if pi["special_conditions"]:
        P.append('<h3 class="sub">특기사항</h3>')
        ul(pi["special_conditions"])
    if pi["unit_program"]:
        P.append('<h3 class="sub">단위세대·시설별 분배</h3>')
        rows = []
        for u in pi["unit_program"]:
            if not isinstance(u, dict):
                rows.append(f"<li>{_esc(_str_item(u))}</li>")
                continue
            head = (u.get("block") or "") + (f"({u.get('tenure')})" if u.get("tenure") else "")
            if u.get("type_label"):
                head = (head + " · " if head else "") + u["type_label"]
            parts = [x for x in (u.get("area_text"), u.get("ratio_text")) if x]
            if u.get("note"):
                parts.append(f"비고: {u['note']}")
            body = " / ".join(parts) if parts else "(내용 없음)"
            rows.append(f"<li>{_esc((head + ': ' if head else '') + body)}</li>")
        if rows:
            P.append('<ul class="list">' + "".join(rows) + "</ul>")

    for i, st in enumerate([x for x in pi["sites"] if isinstance(x, dict)]):
        sid = st.get("site_id") or f"부지{i+1}"
        P.append(f'<h3 class="sub">부지개요 · {_esc(sid)}</h3>')
        fac = st.get("facilities") or []
        kv([
            ("위치", st.get("address")), ("용도지역지구", st.get("zoning")),
            ("공모범위", st.get("scope")),
            ("도입시설", ", ".join(fac) if fac else None),
            ("대지면적", st.get("site_area_sqm")), ("연면적", st.get("floor_area_sqm")),
            ("건폐율", st.get("building_coverage_pct")), ("용적률", st.get("floor_area_ratio_pct")),
            ("최고높이", st.get("max_height_m")), ("공개공지", st.get("open_space_sqm")),
            ("공개공지조건", st.get("open_space_notes")),
        ], units={"대지면적": " ㎡", "연면적": " ㎡", "건폐율": "%", "용적률": "%",
                  "최고높이": " m", "공개공지": " ㎡"})

    for i, st in enumerate([x for x in a["sites"] if isinstance(x, dict)]):
        sid = st.get("site_id") or f"부지{i+1}"
        zoning = st.get("zoning") or []
        fac = st.get("facilities") or []
        P.append(f'<h3 class="sub">건축개요 · {_esc(sid)}</h3>')
        kv([
            ("위치", st.get("address")),
            ("지역지구", ", ".join(zoning) if isinstance(zoning, list) else zoning),
            ("건축구분", st.get("construction_type")), ("건축용도", st.get("building_use")),
            ("도입시설", ", ".join(fac) if fac else None),
            ("대지면적", st.get("site_area_sqm")), ("연면적", st.get("floor_area_sqm")),
            ("건폐율한도", st.get("building_coverage_limit_pct")),
            ("용적률한도", st.get("floor_area_ratio_limit_pct")),
            ("최고높이", st.get("max_height_m")),
            ("공개공지", st.get("public_open_space_sqm")),
            ("공개공지조건", st.get("public_open_space_notes")),
        ], units={"대지면적": " ㎡", "연면적": " ㎡", "건폐율한도": "%", "용적률한도": "%",
                  "최고높이": " m", "공개공지": " ㎡"})

    P.append('<h3 class="sub">전체 규모 한도</h3>')
    kv([
        ("요구연면적", a["total_fa"]), ("대지면적", a["site_area"]),
        ("건폐율한도", a["bcr"]), ("용적률한도", a["far"]), ("높이한도", a["height"]),
        ("지상층수", a["floors_above"]), ("지하층수", a["floors_below"]),
        ("주차대수", a["parking"]),
        ("예정공사비", a["construction_cost"]), ("예정설계비", a["design_fee"]),
        ("설계기간", a["design_period"]),
    ], units={"요구연면적": " ㎡", "대지면적": " ㎡", "건폐율한도": "%", "용적률한도": "%",
              "높이한도": " m", "지상층수": " 층", "지하층수": " 층", "주차대수": " 대"})
    endsec()

    # ══ 2. 면적 프로그램 ════════════════════════════════════════════════════════
    sec(2, "면적 프로그램", "면적")
    _LEVEL = {"site_total": (0, "부지"), "facility": (1, "시설"),
              "bureau": (2, "영역"), "division": (3, "과"), "space": (4, "세부")}

    if a["area_rows"]:
        rows = [ar for ar in a["area_rows"] if isinstance(ar, dict)]

        def _row_html(ar: dict) -> str:
            lvl, lbl = _LEVEL.get(ar.get("row_type") or "space", (4, "세부"))
            name = _esc(ar.get("name") or "")
            area_raw = ar.get("area") if ar.get("area") is not None else ar.get("subtotal_area")
            area_str = _v(area_raw, " ㎡") if area_raw is not None else ""
            meta = []
            if ar.get("notes"):
                meta.append(_esc(ar["notes"]))
            if ar.get("dept"):
                meta.append("소관 " + _esc(ar["dept"]))
            meta_html = f'<span class="meta">{" · ".join(meta)}</span>' if meta else ""
            cls = "row subtotal" if ar.get("is_subtotal") else "row"
            pad = max(0, (lvl - 2)) * 16   # 시설 그룹 내부 기준으로 들여쓰기 재정렬
            return (f'<div class="{cls}"><span class="lv">{_esc(lbl)}</span>'
                    f'<span class="nm" style="padding-left:{pad}px">{name}{meta_html}</span>'
                    f'<span class="ar">{_esc(area_str)}</span></div>')

        def _area_str(ar: dict) -> str:
            raw = ar.get("subtotal_area") if ar.get("subtotal_area") is not None else ar.get("area")
            return _v(raw, " ㎡") if raw is not None else ""

        def _facility_row(ar: dict) -> str:
            """자식 없는 시설(요약 표) — 평범한 행으로 렌더 (빈 아코디언 방지)."""
            return (f'<div class="row subtotal"><span class="lv">시설</span>'
                    f'<span class="nm">{_esc(ar.get("name") or "")}</span>'
                    f'<span class="ar">{_esc(_area_str(ar))}</span></div>')

        # 시설(facility) 단위로 접기 그룹화. 자식 있는 시설만 아코디언, 자식 없는 요약
        # 시설은 평범한 행. site_total 은 독립 합계바. 1000행 벽 방지.
        has_detail = sum(1 for ar in rows
                         if (ar.get("row_type") or "space") not in ("site_total", "facility"))
        collapse = has_detail > 40          # 상세 행이 많으면 기본 접힘
        out: list[str] = []
        buf: list[str] = []                # 현재 시설 그룹의 자식 행
        head: dict | None = None           # 현재 시설 헤더 행
        cnt = 0                            # 현재 그룹 세부(space) 수
        flat: list[str] = []               # 연속된 요약 시설/고아 행 누적

        def _flush_flat():
            nonlocal flat
            if flat:
                out.append('<div class="tree">' + "".join(flat) + "</div>")
                flat = []

        def _flush():
            nonlocal buf, head, cnt
            if head is not None:
                if buf:                    # 자식 있음 → 아코디언
                    _flush_flat()
                    meta = _area_str(head)
                    if cnt:
                        meta = (meta + " · " if meta else "") + f"{cnt}개 항목"
                    openattr = "" if collapse else " open"
                    out.append(
                        f'<details class="fac"{openattr}><summary>'
                        f'<span class="caret">▶</span>'
                        f'<span class="fname">{_esc(head.get("name") or "")}</span>'
                        f'<span class="fmeta">{_esc(meta)}</span></summary>'
                        f'<div class="body"><div class="tree">{"".join(buf)}</div></div></details>'
                    )
                else:                      # 자식 없음 → 요약 행
                    flat.append(_facility_row(head))
            elif buf:                      # 시설 헤더 없는 고아 행
                _flush_flat()
                out.append('<div class="tree">' + "".join(buf) + "</div>")
            buf, head, cnt = [], None, 0

        for ar in rows:
            rt = ar.get("row_type") or "space"
            if rt == "site_total":
                _flush()
                _flush_flat()
                out.append(f'<div class="totalbar"><span>{_esc(ar.get("name") or "")}</span>'
                           f'<span class="ar">{_esc(_area_str(ar))}</span></div>')
            elif rt == "facility":
                _flush()
                head = ar
            else:
                buf.append(_row_html(ar))
                if rt == "space":
                    cnt += 1
        _flush()
        _flush_flat()

        if collapse and out:
            out.insert(0, '<div class="note muted">시설명을 눌러 펼치세요 · '
                          f'세부 {has_detail:,}행</div>')
        P.append("".join(out))

    elif a["area_table"]:
        for grp in a["area_table"]:
            if not isinstance(grp, dict):
                continue
            P.append(f'<h3 class="sub">{_esc(grp.get("group_name") or "(무제)")}</h3>')
            kv([("부지", grp.get("site_id") or "-"),
                ("합계면적", grp.get("total_area_sqm"))], units={"합계면적": " ㎡"})
            trs = []
            for it in (grp.get("items") or []):
                if not isinstance(it, dict):
                    continue
                trs.append(f'<tr><td>{_esc(it.get("name") or "")}</td>'
                           f'<td class="num">{_esc(_v(it.get("area_sqm"), " ㎡"))}</td>'
                           f'<td class="muted">{_esc(it.get("notes") or "")}</td></tr>')
                for sub in (it.get("sub_items") or []):
                    if not isinstance(sub, dict):
                        continue
                    trs.append(f'<tr><td style="padding-left:26px" class="muted">└ {_esc(sub.get("name") or "")}</td>'
                               f'<td class="num">{_esc(_v(sub.get("area_sqm"), " ㎡"))}</td>'
                               f'<td class="muted">{_esc(sub.get("notes") or "")}</td></tr>')
            if trs:
                P.append('<table class="t"><thead><tr><th>항목</th><th class="num">면적</th>'
                         '<th>비고</th></tr></thead><tbody>' + "".join(trs) + "</tbody></table>")
        if a["shared_areas"]:
            P.append('<h3 class="sub">공용·공동 면적</h3>')
            trs = []
            for sa in a["shared_areas"]:
                if not isinstance(sa, dict):
                    continue
                trs.append(f'<tr><td>{_esc(sa.get("name") or "")}</td>'
                           f'<td class="num">{_esc(_v(sa.get("area_sqm"), " ㎡"))}</td>'
                           f'<td class="muted">{_esc(sa.get("notes") or "")}</td></tr>')
            P.append('<table class="t"><thead><tr><th>항목</th><th class="num">면적</th>'
                     '<th>비고</th></tr></thead><tbody>' + "".join(trs) + "</tbody></table>")

    elif a["rooms"]:
        P.append('<h3 class="sub">실별 면적 프로그램</h3>')
        trs = []
        for rm in a["rooms"]:
            if not isinstance(rm, dict):
                continue
            area = rm.get("required_area_sqm") or rm.get("area_sqm")
            cnt = rm.get("required_count") or rm.get("count") or 1
            trs.append(f'<tr><td>{_esc(rm.get("name") or "")}</td>'
                       f'<td class="num">{_esc(_v(area, " ㎡"))}</td>'
                       f'<td class="num">{_esc(cnt)}</td>'
                       f'<td class="muted">{_esc(rm.get("floor") or "")}</td></tr>')
        P.append('<table class="t"><thead><tr><th>실명</th><th class="num">면적</th>'
                 '<th class="num">수량</th><th>위치</th></tr></thead><tbody>'
                 + "".join(trs) + "</tbody></table>")

    if not a["area_table"] and a["zones"]:
        P.append('<h3 class="sub">존 구성</h3>')
        trs = [f'<tr><td>{_esc(z.get("name") or z.get("zone") or "")}</td>'
               f'<td class="num">{_esc(_v(z.get("area_sqm"), " ㎡"))}</td></tr>'
               for z in a["zones"] if isinstance(z, dict)]
        P.append('<table class="t"><thead><tr><th>존</th><th class="num">면적</th>'
                 '</tr></thead><tbody>' + "".join(trs) + "</tbody></table>")
    endsec()

    # ══ 3. 심사기준 ════════════════════════════════════════════════════════════
    sec(3, "심사기준", "심사")
    kv([("총배점", e["total_points"]), ("평가방법", e["eval_method"]),
        ("심사단구성", e["jury"])], units={"총배점": " 점"})

    # 동일 이름 병합 (md 와 동일 로직)
    _merged: list[dict] = []
    _seen: dict[str, int] = {}
    for ev in e["rows"]:
        if not isinstance(ev, dict):
            continue
        name = ev.get("name") or "(항목명 없음)"
        if name in _seen:
            ex = _merged[_seen[name]]
            ep, np_ = ex.get("points"), ev.get("points")
            if isinstance(ep, (int, float)) and isinstance(np_, (int, float)):
                ex["points"] = ep + np_
            elif np_ is not None:
                ex["points"] = np_
            ex.setdefault("sub_items", []).extend(ev.get("sub_items") or [])
            sw = ex.get("shared_with") or []
            sw.extend(x for x in (ev.get("shared_with") or []) if x not in sw)
            ex["shared_with"] = sw
        else:
            _seen[name] = len(_merged)
            _merged.append(dict(ev))

    running = 0.0
    trs = []
    for ev in _merged:
        name = _esc(ev.get("name") or "(항목명 없음)")
        pts = ev.get("points")
        if isinstance(pts, (int, float)):
            running += pts
        shared = ev.get("shared_with") or []
        subs = ev.get("sub_items") or []
        detail = []
        if shared:
            detail.append('<span class="muted">공유배점: ' + _esc(", ".join(shared)) + "</span>")
        if ev.get("description"):
            detail.append(_esc(ev["description"]))
        if subs:
            detail.append('<ul class="list">'
                          + "".join(f"<li>{_esc(_str_item(x))}</li>" for x in subs) + "</ul>")
        detail_html = ("<div>" + "</div><div>".join(detail) + "</div>") if detail else ""
        trs.append(f'<tr><td>{name}{detail_html}</td>'
                   f'<td class="num">{_esc(_v(pts, " 점"))}</td></tr>')
    if trs:
        sum_txt = f"{running:g} 점" if running else "(없음)"
        trs.append(f'<tr class="sumrow"><td>배점 합계</td>'
                   f'<td class="num">{_esc(sum_txt)}</td></tr>')
        warn = ('<div class="note muted">⚠ 배점 합계가 총배점과 크게 다릅니다 — '
                '병합셀 중복 집계 가능성 (확인 필요).</div>' if e["points_sum_warning"] else "")
        P.append('<table class="t"><thead><tr><th>평가항목</th><th class="num">배점</th>'
                 '</tr></thead><tbody>' + "".join(trs) + "</tbody></table>" + warn)

    if e["disqualify"]:
        P.append('<h3 class="sub">실격 요건</h3>')
        ul(e["disqualify"])
    endsec()

    # ══ 4. 요구사항·설계 지침 ═══════════════════════════════════════════════════
    sec(4, "요구사항·설계 지침", "지침")
    if r["requirements"]:
        P.append('<h3 class="sub">평가축별 요구사항</h3>')
        trs = []
        for req in r["requirements"]:
            if not isinstance(req, dict):
                continue
            trs.append(f'<tr><td>{_esc(req.get("axis") or "")}</td>'
                       f'<td class="num">{_esc(_v(req.get("weight_pct"), "%"))}</td>'
                       f'<td>{_esc(req.get("description") or "")}</td></tr>')
        P.append('<table class="t"><thead><tr><th>평가축</th><th class="num">배점비중</th>'
                 '<th>설명</th></tr></thead><tbody>' + "".join(trs) + "</tbody></table>")

    if r["concept"]:
        P.append(f'<div class="note"><b>설계방향</b> · {_esc(r["concept"])}</div>')

    grouped_all = s.get("guidelines_grouped") or []
    if grouped_all:
        facility_specific = [g for g in grouped_all if (g.get("facility_scope") or "전체") != "전체"]
        common_grouped = [g for g in grouped_all if (g.get("facility_scope") or "전체") == "전체"]

        def _html_grouped(rows: list[dict], heading: str) -> None:
            if not rows:
                return
            P.append(f'<h3 class="sub">{_esc(heading)}</h3>')
            order: list[str] = []
            by_fac: dict[str, list[dict]] = {}
            for g in rows:
                fs = (g.get("facility_scope") or "전체").strip() or "전체"
                if fs not in by_fac:
                    order.append(fs)
                    by_fac[fs] = []
                by_fac[fs].append(g)
            for fs in order:
                if fs != "전체":
                    P.append(f'<h4 class="subsub">[{_esc(fs)}]</h4>')
                for g in by_fac[fs]:
                    space = (g.get("space_scope") or "전체").strip() or "전체"
                    cat = (g.get("category") or "기타").strip() or "기타"
                    sec_path = (g.get("section_path") or "").strip()
                    subs = g.get("items_by_sub")
                    if not subs:
                        items = g.get("items") or []
                        if not items:
                            continue
                        subs = [{"sub_path": "", "items": items}]
                    head_parts = []
                    if space != "전체":
                        head_parts.append(f"[{space}]")
                    head_parts.append(cat)
                    if sec_path:
                        head_parts.append(f"— {sec_path}")
                    P.append('<div class="gblock"><div class="ghead">'
                             + _esc(" ".join(head_parts)) + "</div>")
                    for sub in subs:
                        sub_path = (sub.get("sub_path") or "").strip()
                        sub_items = [it for it in (sub.get("items") or [])
                                     if isinstance(it, dict) and (it.get("text") or "").strip()]
                        if not sub_items:
                            continue
                        if sub_path:
                            P.append(f'<div class="gsub">{_esc(sub_path)}</div>')
                        lis = []
                        for it in sub_items:
                            label = (it.get("label") or "").strip()
                            text = (it.get("text") or "").strip()
                            prefix = "" if (not label or label == "-") else _esc(label) + " "
                            lis.append(f"<li>{prefix}{_esc(text)}</li>")
                        P.append('<ul class="list">' + "".join(lis) + "</ul>")
                    P.append("</div>")

        _html_grouped(facility_specific, "시설별 지침")
        _html_grouped(common_grouped, "설계 지침 및 요구사항")

    if not grouped_all:
        _misc = [
            ("특수요구사항", r["special_reqs"]), ("기타설계지침", r["design_reqs"]),
            ("후퇴선요건", r["setbacks"]), ("재료요건", r["materials"]),
            ("친환경요건", r["sustainability"]), ("금지사항", r["prohibited"]),
            ("특별지침", r["special_guide"]),
        ]
        if any(v for _, v in _misc):
            P.append('<h3 class="sub">기타 설계 지침</h3>')
            for lbl, lst in _misc:
                if lst:
                    P.append(f'<h4 class="subsub">{_esc(lbl)}</h4>')
                    ul(lst)
    endsec()

    # ══ 5. 검증 경고 ════════════════════════════════════════════════════════════
    sec(5, "검증 경고", "경고")
    P.append(
        '<div class="warn-sum">'
        f'<span class="pill high"><span class="dot"></span>높음 <b>{high_n}</b></span>'
        f'<span class="pill med"><span class="dot"></span>보통 <b>{medium_n}</b></span>'
        f'<span class="pill"><span class="dot"></span>낮음 <b>{low_n}</b></span></div>'
    )
    if flags:
        ws = []
        for f in flags:
            sv = f.get("severity", "")
            sev = _SEVERITY_LABEL.get(sv, sv)
            loc = f'<div class="wl">위치: {_esc(f.get("location"))}</div>' if f.get("location") else ""
            ws.append(
                f'<div class="w {_esc(sv)}"><div class="wt">{_esc(f.get("type") or "")}'
                f'<span class="sev">{_esc(sev)}</span></div>'
                f'<div class="wm">{_esc(f.get("message") or "")}</div>{loc}</div>'
            )
        P.append('<div class="warn-list">' + "".join(ws) + "</div>")
    else:
        P.append('<div class="note ok">경고 없음 — 검증 통과</div>')
    endsec()

    P.append('<footer class="doc">Competition Analyzer · 지침서 분석 리포트</footer>')

    # 상단 고정 내비게이션 (섹션 점프) — body 본문 위에 배치
    nav_links = "".join(f'<a href="#{anc}">{_esc(lbl)}</a>' for anc, lbl in nav_items)
    nav_html = (
        '<nav class="top"><div class="inner">'
        f'<span class="ttl">{_esc(title)}</span>'
        f'<span class="links">{nav_links}</span></div></nav>'
    )

    return (
        "<!DOCTYPE html><html lang=\"ko\"><head><meta charset=\"utf-8\">"
        "<meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">"
        f"<title>{_esc(title)} · 지침서 분석</title><style>{_HTML_CSS}</style></head>"
        f'<body>{nav_html}<div class="wrap">{"".join(P)}</div></body></html>'
    )


# ── Excel (openpyxl) ──────────────────────────────────────────────────────────

def to_xlsx(brief_data: dict, validation: dict) -> bytes:
    """지침서 체크리스트를 Excel(.xlsx) bytes로 반환.

    시트: 1.면적·프로그램(사업개요 서브섹션 포함) / 2.심사기준 / 3.요구사항 /
    4.검증경고 (+ area_rows 있으면 5.면적표상세).
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl 패키지가 설치되지 않았습니다. "
            "'pip install openpyxl' 후 재시작하세요."
        ) from exc

    s    = _extract_sections(brief_data)
    a, e, r = s["area"], s["eval"], s["reqs"]
    pi = s["project_info"]
    flags = sorted(
        validation.get("flags") or [],
        key=lambda f: _SEVERITY_ORDER.get(f.get("severity", "low"), 2),
    )
    summary = validation.get("summary") or {}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 Sheet 제거

    # ── 공통 스타일 ───────────────────────────────────────────────────────────
    _bold      = Font(bold=True)
    _hdr_font  = Font(bold=True, color="FFFFFF")
    _hdr_fill  = PatternFill("solid", fgColor="3D3D6B")   # 남색 헤더
    _sec_fill  = PatternFill("solid", fgColor="E4E4F0")   # 섹션 제목 배경
    _sub_fill  = PatternFill("solid", fgColor="F2F2FA")   # 소제목 배경
    _grp_fill  = PatternFill("solid", fgColor="DCDCF0")   # area_table 그룹 행 배경
    _center    = Alignment(horizontal="center", vertical="center")
    _wrap_top  = Alignment(wrap_text=True, vertical="top")

    _SEVERITY_FILL = {
        "high":   PatternFill("solid", fgColor="FFCCCC"),
        "medium": PatternFill("solid", fgColor="FFF3CD"),
        "low":    PatternFill("solid", fgColor="D9EDF7"),
    }
    _warn_fill   = PatternFill("solid", fgColor="FF9900")  # 배점 합계 경고
    _thin        = Side(style="thin")
    _border_thin = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _num_right   = Alignment(horizontal="right", vertical="top")
    _NUM_FMT     = "#,##0.##"

    def _sep(ws, r: int) -> int:
        """구분 빈 행 높이를 8pt로 설정하고 다음 행 번호 반환."""
        ws.row_dimensions[r].height = 8
        return r + 1

    def _write_section_title(ws, title: str, row: int, span: int = 4) -> int:
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(bold=True, size=13, color="1A1A4E")
        c.fill = _sec_fill
        c.alignment = _center
        if span > 1:
            ws.merge_cells(
                start_row=row, start_column=1,
                end_row=row, end_column=span,
            )
        ws.row_dimensions[row].height = 22
        return row + 1

    def _write_subsection(ws, title: str, row: int, span: int = 4) -> int:
        c = ws.cell(row=row, column=1, value=title)
        c.font = _bold
        c.fill = _sub_fill
        if span > 1:
            ws.merge_cells(
                start_row=row, start_column=1,
                end_row=row, end_column=span,
            )
        return row + 1

    def _write_header(ws, cols: list[str], row: int) -> int:
        for ci, label in enumerate(cols, 1):
            c = ws.cell(row=row, column=ci, value=label)
            c.font = _hdr_font
            c.fill = _hdr_fill
            c.alignment = _center
        return row + 1

    def _write_kv(ws, label: str, val: Any, row: int, val_end_col: int = 2, num_format: str = "") -> int:
        ws.cell(row=row, column=1, value=label).font = _bold
        raw = _cell_safe(val)
        c = ws.cell(row=row, column=2, value=raw)
        if num_format and isinstance(raw, (int, float)):
            c.number_format = num_format
            c.alignment = _num_right
        else:
            c.alignment = _wrap_top
        if val_end_col > 2:
            ws.merge_cells(start_row=row, start_column=2,
                           end_row=row, end_column=val_end_col)
        return row + 1

    def _auto_width(ws) -> None:
        """셀 최대 길이 기준 열 너비 자동 조정 (최소 10, 최대 60)."""
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=6,
            )
            ws.column_dimensions[
                get_column_letter(col[0].column)
            ].width = max(min(max_len + 4, 60), 10)

    # ── Sheet 1: 면적·프로그램 ────────────────────────────────────────────────
    ws1 = wb.create_sheet("1.면적·프로그램")

    _SITE_COLS = ["부지ID", "위치", "지역지구", "건축구분", "건축용도", "도입시설",
                  "대지면적(㎡)", "연면적(㎡)", "건폐율(%)", "용적률(%)", "높이(m)",
                  "공개공지(㎡)", "공개공지 조건"]
    multi_sites = [s for s in a["sites"] if isinstance(s, dict)] if len(a["sites"]) > 1 else []
    bpi_sites_xl = [st for st in pi["sites"] if isinstance(st, dict)]
    has_bpi_xl = bool(bpi_sites_xl or pi["competition_name"] or pi["construction_cost_100m_won"] is not None)

    # BPI 부지는 KV 블록(4열)으로 표시 — BRIEF_PROGRAM 복수부지 테이블이 있으면 그 열수 사용
    _span1 = max(len(_SITE_COLS) if multi_sites else 0, 4)
    row = _write_section_title(ws1, "1. 면적·프로그램 요구", 1, span=_span1)
    row += 1

    # ── 사업 개요 (BRIEF_PROJECT_INFO) ────────────────────────────────────────
    ws1.freeze_panes = "A3"
    ws1.print_title_rows = "1:2"

    if has_bpi_xl:
        row = _write_subsection(ws1, "사업 개요", row, span=_span1)
        if pi["competition_name"]:
            row = _write_kv(ws1, "공모명", pi["competition_name"], row, val_end_col=4)
            ws1.cell(row=row - 1, column=2).font = Font(bold=True, size=12)
        for label, val in [
            ("발주처",   pi["organizer"] or None),
            ("공모유형", pi["competition_type"] or None),
        ]:
            if val:
                row = _write_kv(ws1, label, val, row, val_end_col=4)
        for i, st in enumerate(bpi_sites_xl):
            sid = st.get("site_id") or f"부지{i+1}"
            row = _write_subsection(ws1, sid, row, span=4)
            for label, val in [
                ("위치",          st.get("address") or None),
                ("용도지역/지구", st.get("zoning") or None),
                ("공모범위",      st.get("scope") or None),
                ("도입시설",      ", ".join(st.get("facilities") or []) or None),
                ("대지면적(㎡)", _cell_safe(st.get("site_area_sqm"))),
                ("연면적(㎡)",   _cell_safe(st.get("floor_area_sqm"))),
                ("건폐율(%)",    _cell_safe(st.get("building_coverage_pct"))),
                ("용적률(%)",    _cell_safe(st.get("floor_area_ratio_pct"))),
                ("최고높이(m)",  _cell_safe(st.get("max_height_m"))),
                ("공개공지(㎡)", _cell_safe(st.get("open_space_sqm"))),
                ("공개공지 조건", st.get("open_space_notes") or None),
            ]:
                if val is not None and val != "":
                    row = _write_kv(ws1, label, val, row, val_end_col=4)
            row += 1
        for label, val in [
            ("예정 공사비 (억원)", pi["construction_cost_100m_won"]),
            ("예정 설계비 (억원)", pi["design_cost_100m_won"]),
            ("공사 기간 (개월)",  pi["construction_period_months"]),
        ]:
            if val is not None:
                row = _write_kv(ws1, label, val, row, val_end_col=4, num_format=_NUM_FMT)
        for title, items in [
            ("예산 산정 기준", pi["budget_notes"]),
            ("특기사항",       pi["special_conditions"]),
        ]:
            if items:
                row = _write_subsection(ws1, title, row, span=4)
                for item in items:
                    c = ws1.cell(row=row, column=1, value=f"• {_str_item(item)}")
                    c.alignment = _wrap_top
                    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                    row += 1

        # 단위세대·시설별 분배표 (BRIEF_PROJECT_INFO.unit_program[])
        unit_prog = [u for u in pi.get("unit_program") or [] if isinstance(u, dict)]
        if unit_prog:
            row = _write_subsection(ws1, "단위세대·시설별 분배", row, span=4)
            row = _write_header(ws1, ["구분", "평형/유형", "면적·규모", "비율/비고"], row)
            for u in unit_prog:
                block      = u.get("block") or ""
                tenure     = u.get("tenure") or ""
                type_label = u.get("type_label") or ""
                area_text  = u.get("area_text") or ""
                ratio_text = u.get("ratio_text") or ""
                note       = u.get("note") or ""
                head = block + (f"({tenure})" if tenure else "")
                # 비율/비고: 비율과 비고를 한 칸에 결합 (둘 다 있을 때는 줄바꿈)
                rn_parts: list[str] = []
                if ratio_text:
                    rn_parts.append(ratio_text)
                if note:
                    rn_parts.append(note)
                rn_combined = "\n".join(rn_parts)
                for c_idx, v in enumerate([head, type_label, area_text, rn_combined], start=1):
                    cell = ws1.cell(row=row, column=c_idx, value=v or "")
                    cell.alignment = _wrap_top
                row += 1
        row = _sep(ws1, row)

    # ── 전체 규모 한도 / 부지별 건축개요 ─────────────────────────────────────
    if multi_sites:
        # 복수 부지: 부지별 상세 테이블
        row = _write_subsection(ws1, "부지별 건축개요", row, span=len(_SITE_COLS))
        row = _write_header(ws1, _SITE_COLS, row)
        for i, st in enumerate(multi_sites):
            zoning_str = ", ".join(st.get("zoning") or []) or ""
            fac_str    = ", ".join(st.get("facilities") or []) or ""
            vals = [
                st.get("site_id") or f"부지{i+1}",
                st.get("address") or "",
                zoning_str,
                st.get("construction_type") or "",
                st.get("building_use") or "",
                fac_str,
                _cell_safe(st.get("site_area_sqm")),
                _cell_safe(st.get("floor_area_sqm")),
                _cell_safe(st.get("building_coverage_limit_pct")),
                _cell_safe(st.get("floor_area_ratio_limit_pct")),
                _cell_safe(st.get("max_height_m")),
                _cell_safe(st.get("public_open_space_sqm")),
                st.get("public_open_space_notes") or "",
            ]
            for ci, v in enumerate(vals, 1):
                c = ws1.cell(row=row, column=ci, value=v)
                c.alignment = _wrap_top
            row += 1
        row = _sep(ws1, row)
        row = _write_subsection(ws1, "공통 규모", row, span=4)
        for label, val in [
            ("요구 총 연면적 (㎡)", a["total_fa"]),
            ("건폐율 한도 (%)",     a["bcr"]),
            ("용적률 한도 (%)",     a["far"]),
            ("지상 층수",           a["floors_above"]),
            ("지하 층수",           a["floors_below"]),
            ("주차 대수 (대)",      a["parking"]),
            ("예정 공사비",         a["construction_cost"] or None),
            ("예정 설계비",         a["design_fee"] or None),
            ("설계 기간",           a["design_period"] or None),
        ]:
            row = _write_kv(ws1, label, val, row, val_end_col=4, num_format=_NUM_FMT)
    else:
        row = _write_subsection(ws1, "전체 규모 한도", row, span=4)
        for label, val in [
            ("대지면적 (㎡)",    a["site_area"]),
            ("요구 연면적 (㎡)", a["total_fa"]),
            ("건폐율 한도 (%)",  a["bcr"]),
            ("용적률 한도 (%)",  a["far"]),
            ("높이 한도 (m)",    a["height"]),
            ("지상 층수",        a["floors_above"]),
            ("지하 층수",        a["floors_below"]),
            ("주차 대수 (대)",   a["parking"]),
            ("예정 공사비",      a["construction_cost"] or None),
            ("예정 설계비",      a["design_fee"] or None),
            ("설계 기간",        a["design_period"] or None),
        ]:
            row = _write_kv(ws1, label, val, row, val_end_col=4, num_format=_NUM_FMT)
    row = _sep(ws1, row)

    _AT_COLS = ["구분", "기준면적(A)", "계획면적(B)", "비고"]

    if a["area_rows"]:
        # 상세 면적표는 Sheet 5에 렌더링 — 여기엔 참조 안내만 표시
        c_note = ws1.cell(row=row, column=1,
                          value="※ 상세 면적 프로그램은 시트 5 「면적표상세」 참조")
        c_note.font = Font(italic=True, color="444444")
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
    elif a["area_table"]:
        row = _write_subsection(ws1, "실별 면적 프로그램", row, span=4)
        row = _write_header(ws1, _AT_COLS, row)
        for grp in a["area_table"]:
            if not isinstance(grp, dict):
                continue
            # 그룹 행: 굵게 + _grp_fill + 테두리
            c = ws1.cell(row=row, column=1, value=grp.get("group_name") or "")
            c.font = _bold; c.fill = _grp_fill; c.border = _border_thin
            c2 = ws1.cell(row=row, column=2, value=grp.get("total_area_sqm"))
            c2.font = _bold; c2.fill = _grp_fill; c2.alignment = _num_right
            c2.number_format = _NUM_FMT; c2.border = _border_thin
            for ci in (3, 4):
                ct = ws1.cell(row=row, column=ci)
                ct.fill = _grp_fill; ct.border = _border_thin
            row += 1
            for item in (grp.get("items") or []):
                if not isinstance(item, dict):
                    continue
                c = ws1.cell(row=row, column=1, value=item.get("name") or "")
                c.alignment = Alignment(indent=2, vertical="top")
                c.border = _border_thin
                c2 = ws1.cell(row=row, column=2, value=item.get("area_sqm"))
                c2.alignment = _num_right; c2.number_format = _NUM_FMT; c2.border = _border_thin
                # column 3 = 계획면적(B) — 빈칸 (설계자 입력용)
                c3 = ws1.cell(row=row, column=3)
                c3.alignment = _num_right; c3.number_format = _NUM_FMT; c3.border = _border_thin
                c4 = ws1.cell(row=row, column=4, value=item.get("notes") or "")
                c4.border = _border_thin
                row += 1
                for sub in (item.get("sub_items") or []):
                    if not isinstance(sub, dict):
                        continue
                    c = ws1.cell(row=row, column=1, value=sub.get("name") or "")
                    c.alignment = Alignment(indent=4, vertical="top")
                    c.border = _border_thin
                    c2 = ws1.cell(row=row, column=2, value=sub.get("area_sqm"))
                    c2.alignment = _num_right; c2.number_format = _NUM_FMT; c2.border = _border_thin
                    c3 = ws1.cell(row=row, column=3)
                    c3.alignment = _num_right; c3.number_format = _NUM_FMT; c3.border = _border_thin
                    c4 = ws1.cell(row=row, column=4, value=sub.get("notes") or "")
                    c4.border = _border_thin
                    row += 1
        row = _sep(ws1, row)
        if a["shared_areas"]:
            row = _write_subsection(ws1, "공용·공동 면적", row, span=4)
            row = _write_header(ws1, _AT_COLS, row)
            for sa in a["shared_areas"]:
                if not isinstance(sa, dict):
                    continue
                c1 = ws1.cell(row=row, column=1, value=sa.get("name") or "")
                c1.border = _border_thin
                c2 = ws1.cell(row=row, column=2, value=sa.get("area_sqm"))
                c2.alignment = _num_right; c2.number_format = _NUM_FMT; c2.border = _border_thin
                c3 = ws1.cell(row=row, column=3); c3.border = _border_thin
                c4 = ws1.cell(row=row, column=4, value=sa.get("notes") or "")
                c4.border = _border_thin
                row += 1
            row = _sep(ws1, row)
    elif a["rooms"]:
        # 구 경로 폴백
        row = _write_subsection(ws1, "실별 면적 프로그램", row, span=5)
        row = _write_header(ws1, ["실명", "요구면적(㎡)", "개수", "위치/층", "비고"], row)
        for rm in a["rooms"]:
            if not isinstance(rm, dict):
                ws1.cell(row=row, column=1, value=_str_item(rm))
                row += 1
                continue
            ws1.cell(row=row, column=1, value=rm.get("name") or "")
            area = rm.get("required_area_sqm") or rm.get("area_sqm")
            ws1.cell(row=row, column=2, value=area)
            ws1.cell(row=row, column=3, value=rm.get("required_count") or rm.get("count") or 1)
            ws1.cell(row=row, column=4, value=rm.get("floor") or "")
            ws1.cell(row=row, column=5, value=rm.get("notes") or "")
            row += 1
        row = _sep(ws1, row)

    if not a["area_rows"] and not a["area_table"] and a["zones"]:
        row = _write_subsection(ws1, "존 구성", row, span=2)
        row = _write_header(ws1, ["존명", "면적(㎡)"], row)
        for z in a["zones"]:
            ws1.cell(row=row, column=1, value=z.get("name") or z.get("zone") or "")
            ws1.cell(row=row, column=2, value=z.get("area_sqm"))
            row += 1

    # Sheet 1 열 너비: A(레이블)=16, B(값)=55 고정 — 면적표·KV 블록 기준
    ws1.column_dimensions["A"].width = 16
    ws1.column_dimensions["B"].width = 55
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 14
    # BRIEF_PROGRAM 복수부지 테이블(E열 이후)만 자동 조정
    for col in ws1.columns:
        letter = get_column_letter(col[0].column)
        if letter in ("A", "B", "C", "D"):
            continue
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=6,
        )
        ws1.column_dimensions[letter].width = max(min(max_len + 4, 40), 10)

    # ── Sheet 2: 심사기준 ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("2.심사기준")
    _S2_SPAN = 3  # 구분 | 세부기준 | 배점
    row = _write_section_title(ws2, "2. 심사기준 (배점표)", 1, span=_S2_SPAN)
    row += 1
    ws2.freeze_panes = "A3"
    ws2.print_title_rows = "1:2"

    for label, val in [
        ("총 배점",    e["total_points"]),
        ("평가 방법",  e["eval_method"] or None),
        ("심사단 구성", e["jury"] or None),
    ]:
        row = _write_kv(ws2, label, val, row, val_end_col=_S2_SPAN, num_format=_NUM_FMT)
    row = _sep(ws2, row)

    if e["rows"]:
        row = _write_header(ws2, ["구분", "세부기준", "배점"], row)
        running_total: float = 0.0

        def _group_by_name(rows: list) -> list:
            """같은 이름의 연속 항목을 하나의 그룹으로 묶음 (Col A 병합 기준)."""
            groups: list = []
            for ev in rows:
                if not isinstance(ev, dict):
                    groups.append([ev]); continue
                name = ev.get("name") or ""
                if (groups
                        and isinstance(groups[-1][-1], dict)
                        and (groups[-1][-1].get("name") or "") == name):
                    groups[-1].append(ev)
                else:
                    groups.append([ev])
            return groups

        def _share_runs(rows: list) -> list[tuple[int, int, float | int | None]]:
            """배점 공유 (shared_with) 가 있는 연속 행 구간을 탐지.

            반환: [(start_index, end_index_inclusive, shared_points), ...]
            - PDF 의 병합 셀: A행 points=40, shared_with=["B"] / B행 points=null, shared_with=["A"]
              → 연속 행이고 서로 mutual reference 면 한 구간으로 묶음.
            - shared_with 가 빈 항목 or 비-dict 는 단일 행 구간.
            - 한 구간 안에서 non-null points 1개 → 그게 그룹 배점.
            - Fallback: shared_with 가 비어 있어도 다음 행 points=None 이고 현재 run 에
              이미 non-null points 가 있으면 같은 구간으로 흡수 (LLM 이 vMerge 쌍의
              shared_with 를 누락하는 케이스 대비).
            """
            runs: list[tuple[int, int, float | int | None]] = []
            i = 0
            n = len(rows)
            while i < n:
                ev = rows[i]
                if not isinstance(ev, dict):
                    runs.append((i, i, None))
                    i += 1
                    continue
                cur_name = (ev.get("name") or "").strip()
                cur_sw = [s.strip() for s in (ev.get("shared_with") or []) if s and isinstance(s, str)]
                # 다음 행들 중 shared_with 사이클로 묶인 연속 구간 탐색
                j = i
                names_in_run = {cur_name}
                sw_targets = set(cur_sw)
                while j + 1 < n:
                    nxt = rows[j + 1]
                    if not isinstance(nxt, dict):
                        break
                    nxt_name = (nxt.get("name") or "").strip()
                    nxt_sw = [s.strip() for s in (nxt.get("shared_with") or []) if s and isinstance(s, str)]
                    nxt_pts = nxt.get("points")
                    # 현재 run 에 non-null points 가 이미 있는지 (fallback 매칭 조건)
                    run_has_pts = any(
                        isinstance(rows[k], dict) and rows[k].get("points") is not None
                        for k in range(i, j + 1)
                    )
                    # 다음 행이 현재 그룹 안 누군가를 참조하거나, 현재 그룹이 다음 행을 참조하면 같은 구간
                    # Fallback: 다음 행 points=None 이고 현재 run 에 non-null points 있으면 흡수
                    if (nxt_name and nxt_name in sw_targets) \
                            or any(s in names_in_run for s in nxt_sw) \
                            or (nxt_pts is None and run_has_pts):
                        names_in_run.add(nxt_name)
                        sw_targets.update(nxt_sw)
                        j += 1
                    else:
                        break
                # 구간 (i..j) 의 points 결정: non-null 1개 picks; 여러 개면 첫째 사용
                shared_pts = None
                for k in range(i, j + 1):
                    r_ev = rows[k]
                    if isinstance(r_ev, dict) and r_ev.get("points") is not None:
                        shared_pts = r_ev.get("points")
                        break
                runs.append((i, j, shared_pts))
                i = j + 1
            return runs

        def _fmt_bullets(subs: list, desc: str) -> str:
            """sub_items 리스트를 셀 내 멀티라인 불릿 텍스트로 변환.
            - 복수 항목: 각 항목 한 줄
            - 단일 항목: 개행 → 줄 분리, 인라인 ▪/•/· 구분자 → 줄 분리
            """
            import re as _r
            _BULLET_CHARS = ("▪", "•", "·", "◦", "▸", "▶", "▷")

            def _normalise(t: str) -> str:
                t = t.strip()
                if not t:
                    return ""
                return t if t[0] in _BULLET_CHARS else f"• {t}"

            items = subs if subs else ([desc] if desc else [])
            if not items:
                return ""

            # 복수 sub_items → 각각 한 줄
            if len(items) > 1:
                return "\n".join(_normalise(_str_item(s)) for s in items if _str_item(s).strip())

            text = _str_item(items[0])

            # 개행이 있으면 줄 단위로 분리 (table_rows_raw 원본 텍스트 경로)
            if "\n" in text:
                parts = [p.strip() for p in text.split("\n") if p.strip()]
                if len(parts) > 1:
                    return "\n".join(_normalise(p) for p in parts)

            # 인라인 불릿 문자(▪ • ·)로 분리
            parts = [p.strip() for p in _r.split(r'[▪•·◦]\s*', text) if p.strip()]
            if len(parts) > 1:
                return "\n".join(f"• {p}" for p in parts)

            return text if (text and text[0] in _BULLET_CHARS) else (f"• {text}" if text else "")

        # ── 배점 공유 구간 사전 계산 ─────────────────────────────────────
        # rows 의 절대 index → (run_start, run_end, shared_points) 매핑
        # (주의: for-loop 변수가 enclosing 스코프로 누출되므로 `s`/`t` 같은
        # 짧은 이름 피하기 — Sheet 3 의 `s = sections` 를 덮어쓰면 크래시)
        share_runs = _share_runs(e["rows"])
        share_lookup: dict[int, tuple[int, int, float | int | None]] = {}
        for _rs, _re, _rpts in share_runs:
            for _k in range(_rs, _re + 1):
                share_lookup[_k] = (_rs, _re, _rpts)

        # e["rows"] index → xlsx row 매핑 (사후 Col C 병합 위해)
        abs_to_xlsx_row: dict[int, int] = {}
        # 구간당 1회만 running_total 누적
        counted_runs: set[int] = set()
        row_index = 0  # e["rows"] 누적 index

        for name_group in _group_by_name(e["rows"]):
            n_in_group = len(name_group)
            dicts = [ev for ev in name_group if isinstance(ev, dict)]

            # 비-dict 항목: 전 열 병합 단일 행
            if not dicts:
                for offset, ev in enumerate(name_group):
                    c = ws2.cell(row=row, column=1, value=_str_item(ev))
                    ws2.merge_cells(start_row=row, start_column=1,
                                    end_row=row, end_column=_S2_SPAN)
                    abs_to_xlsx_row[row_index + offset] = row
                    row += 1
                row_index += n_in_group
                continue

            name   = dicts[0].get("name") or ""
            n_rows = len(dicts)

            # ── Col A (구분): 같은 이름 전체 행 병합 ──────────────────
            c1 = ws2.cell(row=row, column=1, value=name)
            c1.font = _bold
            c1.fill = _grp_fill
            c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c1.border = _border_thin
            if n_rows > 1:
                ws2.merge_cells(start_row=row, start_column=1,
                                end_row=row + n_rows - 1, end_column=1)

            for i, ev in enumerate(dicts):
                cur  = row + i
                subs = ev.get("sub_items") or []
                desc = ev.get("description") or ""

                # name_group 안 ev 의 위치 (비-dict 가 섞여 있을 때 안전)
                local_pos = next((k for k, x in enumerate(name_group) if x is ev), i)
                abs_idx = row_index + local_pos
                abs_to_xlsx_row[abs_idx] = cur

                # ── Col B (세부기준) ──────────────────────────────────
                c2 = ws2.cell(row=cur, column=2, value=_fmt_bullets(subs, desc))
                c2.alignment = _wrap_top
                c2.border = _border_thin

                # ── Col C (배점): share run 안 → 첫 행에만 값, 나머지 비움
                run_info = share_lookup.get(abs_idx)
                if run_info is not None:
                    run_start, _, shared_pts = run_info
                    pts = shared_pts if abs_idx == run_start else None
                    # running_total: 첫 등장 시에만 누적
                    if abs_idx == run_start and run_start not in counted_runs:
                        counted_runs.add(run_start)
                        if isinstance(shared_pts, (int, float)):
                            running_total += shared_pts
                else:
                    pts = ev.get("points")
                    if isinstance(pts, (int, float)):
                        running_total += pts

                c3 = ws2.cell(row=cur, column=3, value=pts)
                if pts is not None:
                    c3.font = _bold
                    c3.fill = _grp_fill
                    if isinstance(pts, (int, float)):
                        c3.number_format = _NUM_FMT
                c3.alignment = _center
                c3.border = _border_thin

            row += n_rows
            row_index += n_in_group

        # ── Sheet 2 끝나기 전 post-pass: share run 구간 Col C 병합 ──────────
        # 연속 xlsx 행이면서 같은 share_run 에 속하는 셀들을 한 번에 merge.
        for run_start, run_end, _shared_pts in share_runs:
            if run_end <= run_start:
                continue
            # 구간 안 모든 abs_idx 가 xlsx row 에 매핑되어 있고 연속인지 확인
            xlsx_rows = [abs_to_xlsx_row.get(k) for k in range(run_start, run_end + 1)]
            if any(r is None for r in xlsx_rows):
                continue
            # 연속성 체크
            if all(xlsx_rows[i + 1] == xlsx_rows[i] + 1 for i in range(len(xlsx_rows) - 1)):
                ws2.merge_cells(start_row=xlsx_rows[0], start_column=3,
                                end_row=xlsx_rows[-1], end_column=3)

        # 합계 행 — points_sum_warning 시 주황색 경고 강조
        _sum_fill = _warn_fill if e["points_sum_warning"] else _grp_fill
        c_tot = ws2.cell(row=row, column=1, value="합  계")
        c_tot.font = _bold; c_tot.fill = _sum_fill; c_tot.border = _border_thin
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c_sum = ws2.cell(row=row, column=3,
                         value=running_total if running_total else None)
        c_sum.font = _bold; c_sum.fill = _sum_fill
        c_sum.alignment = _center; c_sum.number_format = _NUM_FMT
        c_sum.border = _border_thin
        row += 2

    if e["disqualify"]:
        row = _write_subsection(ws2, "실격 요건", row, span=_S2_SPAN)
        for d in e["disqualify"]:
            c = ws2.cell(row=row, column=1, value=_str_item(d))
            c.alignment = _wrap_top
            ws2.merge_cells(start_row=row, start_column=1,
                            end_row=row, end_column=_S2_SPAN)
            row += 1

    ws2.column_dimensions["A"].width = 20  # 구분명
    ws2.column_dimensions["B"].width = 55  # 세부기준
    ws2.column_dimensions["C"].width = 10  # 배점

    # ── Sheet 3: 요구사항·필수조건 ────────────────────────────────────────────
    ws3 = wb.create_sheet("3.요구사항")
    _S3_SPAN = 3  # 소항목 레이블 | 내용(col2-3 병합)
    row = _write_section_title(ws3, "3. 요구사항·필수조건", 1, span=_S3_SPAN)
    row += 1
    ws3.freeze_panes = "A3"
    ws3.print_title_rows = "1:2"

    if r["requirements"]:
        row = _write_subsection(ws3, "평가축별 요구사항", row, span=_S3_SPAN)
        row = _write_header(ws3, ["평가축", "설명", "배점비중(%)"], row)
        for req in r["requirements"]:
            if not isinstance(req, dict):
                ws3.cell(row=row, column=1, value=_str_item(req))
                row += 1
                continue
            ws3.cell(row=row, column=1, value=req.get("axis") or "")
            d_cell = ws3.cell(row=row, column=2, value=req.get("description") or "")
            d_cell.alignment = _wrap_top
            ws3.cell(row=row, column=3, value=req.get("weight_pct"))
            row += 1
        row += 1

    if r["concept"]:
        ws3.cell(row=row, column=1, value="설계 방향").font = _bold
        c = ws3.cell(row=row, column=2, value=r["concept"])
        c.alignment = _wrap_top
        ws3.merge_cells(start_row=row, start_column=2,
                        end_row=row, end_column=_S3_SPAN)
        row += 2

    def _ws3_labeled(lbl: str, items: list) -> None:
        """소항목 레이블(col 1) + 내용 불릿(col 2-3 병합) 형태로 출력."""
        nonlocal row
        if not items:
            return
        for i, item in enumerate(items):
            if i == 0:
                c_lbl = ws3.cell(row=row, column=1, value=lbl)
                c_lbl.font = Font(bold=True, size=9)
            text = _str_item(item)
            if not text.startswith("•"):
                text = f"• {text}"
            c = ws3.cell(row=row, column=2, value=text)
            c.alignment = _wrap_top
            ws3.merge_cells(start_row=row, start_column=2,
                            end_row=row, end_column=_S3_SPAN)
            row += 1

    def _ws3_kv(lbl: str, val: Any) -> None:
        """소항목 레이블(col 1) + 단일 값(col 2-3 병합) 형태로 출력."""
        nonlocal row
        if val is None or val == "":
            return
        c_lbl = ws3.cell(row=row, column=1, value=lbl)
        c_lbl.font = Font(bold=True, size=9)
        c = ws3.cell(row=row, column=2, value=_cell_safe(val))
        c.alignment = _wrap_top
        ws3.merge_cells(start_row=row, start_column=2,
                        end_row=row, end_column=_S3_SPAN)
        row += 1

    def _ws3_bullets(items: list) -> None:
        """레이블 없는 평면 불릿 (구 데이터 폴백용)."""
        nonlocal row
        for item in items:
            c = ws3.cell(row=row, column=1, value=f"• {_str_item(item)}")
            c.alignment = _wrap_top
            ws3.merge_cells(start_row=row, start_column=1,
                            end_row=row, end_column=_S3_SPAN)
            row += 1

    # ── 계층 보존 설계지침 (design_guidelines_grouped) ─────────────────────────
    # 시설별 (facility_scope != "전체") 과 공통 (facility_scope == "전체") 으로 분리,
    # facility → space → category 순으로 중첩 렌더링.
    grouped_all = s.get("guidelines_grouped") or []

    def _ws3_grouped_section(rows: list[dict]) -> None:
        """grouped rows 를 facility / space / category 순으로 정렬·렌더."""
        nonlocal row
        # facility_scope 별 그룹화 (순서 유지)
        facility_order: list[str] = []
        by_facility: dict[str, list[dict]] = {}
        for g in rows:
            fs = (g.get("facility_scope") or "전체").strip() or "전체"
            if fs not in by_facility:
                facility_order.append(fs)
                by_facility[fs] = []
            by_facility[fs].append(g)
        for fs in facility_order:
            # 시설 헤더 (멀티 시설일 때만 별도 표시 — facility_scope == "전체" 면 생략)
            if fs != "전체":
                row = _write_subsection(ws3, f"[{fs}]", row, span=_S3_SPAN)
            # space_scope → category 순으로 정렬
            for g in by_facility[fs]:
                space    = (g.get("space_scope") or "전체").strip() or "전체"
                cat      = (g.get("category") or "기타").strip() or "기타"
                sec_path = (g.get("section_path") or "").strip()
                # 정규화 후: items_by_sub 가 권위. 폴백으로 items 처리.
                subs = g.get("items_by_sub")
                if not subs:
                    items = g.get("items") or []
                    if not items:
                        continue
                    subs = [{"sub_path": "", "items": items}]
                else:
                    # 빈 sub 만 있으면 스킵
                    if not any((sub.get("items") or []) for sub in subs):
                        continue
                # 그룹 헤더 (굵게, 한 번만) — "[공간] 카테고리 — section_path"
                head_parts = []
                if space != "전체":
                    head_parts.append(f"[{space}]")
                head_parts.append(cat)
                if sec_path:
                    head_parts.append(f"— {sec_path}")
                head = " ".join(head_parts)
                c_h = ws3.cell(row=row, column=1, value=head)
                c_h.font = _bold
                c_h.alignment = _wrap_top
                ws3.merge_cells(start_row=row, start_column=1,
                                end_row=row, end_column=_S3_SPAN)
                row += 1
                # sub_path 별로 inline sub-header 줄 (굵지 않음) + items
                for sub in subs:
                    sub_path = (sub.get("sub_path") or "").strip()
                    sub_items = sub.get("items") or []
                    if not sub_items:
                        continue
                    if sub_path:
                        # inline sub-header: "- 비품창고" — column 1, 굵지 않음
                        c_sub = ws3.cell(row=row, column=1, value=f"- {sub_path}")
                        c_sub.alignment = _wrap_top
                        ws3.merge_cells(start_row=row, start_column=1,
                                        end_row=row, end_column=_S3_SPAN)
                        row += 1
                    for it in sub_items:
                        if not isinstance(it, dict):
                            continue
                        label = (it.get("label") or "").strip()
                        text  = (it.get("text") or "").strip()
                        if not text:
                            continue
                        line = f"{label} {text}" if label else f"• {text}"
                        c = ws3.cell(row=row, column=2, value=line)
                        c.alignment = _wrap_top
                        ws3.merge_cells(start_row=row, start_column=2,
                                        end_row=row, end_column=_S3_SPAN)
                        row += 1
            row += 1

    if grouped_all:
        facility_specific = [g for g in grouped_all
                             if (g.get("facility_scope") or "전체") != "전체"]
        common_grouped   = [g for g in grouped_all
                             if (g.get("facility_scope") or "전체") == "전체"]
        if facility_specific:
            row = _write_subsection(ws3, "시설별 지침", row, span=_S3_SPAN)
            _ws3_grouped_section(facility_specific)
        if common_grouped:
            row = _write_subsection(ws3, "설계 지침 및 요구사항", row, span=_S3_SPAN)
            _ws3_grouped_section(common_grouped)

    # ── 기타 설계 지침 (폴백 / 구 데이터 — grouped 없을 때만) ─────────────────
    if not grouped_all:
        list_sections = [
            ("특수 요구사항", r["special_reqs"]),
            ("기타 설계 지침", r["design_reqs"]),
            ("후퇴선 요건",   r["setbacks"]),
            ("재료 요건",     r["materials"]),
            ("친환경 요건",   r["sustainability"]),
            ("금지 사항",     r["prohibited"]),
            ("특별 지침",     r["special_guide"]),
        ]
        for lbl, items in list_sections:
            if not items:
                continue
            row = _write_subsection(ws3, lbl, row, span=_S3_SPAN)
            _ws3_bullets(items)
            row += 1

    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 55
    ws3.column_dimensions["C"].width = 12

    # ── Sheet 4: 검증 경고 ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("4.검증경고")
    row = _write_section_title(ws4, "4. 검증 경고", 1, span=4)
    row += 1
    ws4.freeze_panes = "A3"
    ws4.print_title_rows = "1:2"

    high_n, med_n, low_n = (
        summary.get("high", 0),
        summary.get("medium", 0),
        summary.get("low", 0),
    )
    summary_cell = ws4.cell(
        row=row, column=1,
        value=f"요약: 높음 {high_n}건 / 보통 {med_n}건 / 낮음 {low_n}건",
    )
    summary_cell.font = _bold
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    row = _write_header(ws4, ["심각도", "유형", "메시지", "위치"], row)

    if flags:
        for flag in flags:
            sev  = flag.get("severity", "low")
            fill = _SEVERITY_FILL.get(sev, _SEVERITY_FILL["low"])
            vals = [
                _SEVERITY_LABEL.get(sev, sev),
                flag.get("type") or "",
                flag.get("message") or "",
                flag.get("location") or "",
            ]
            for ci, val in enumerate(vals, 1):
                c = ws4.cell(row=row, column=ci, value=val)
                c.fill = fill
                c.alignment = _wrap_top
            row += 1
    else:
        c = ws4.cell(row=row, column=1, value="검출된 경고 없음")
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    ws4.column_dimensions["A"].width = 10
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 60
    ws4.column_dimensions["D"].width = 20

    # ── Sheet 5: 면적표 상세 (area_rows flat 방식) ─────────────────────────────
    _ROW_TYPE_LEVEL = {"site_total": 0, "facility": 1, "bureau": 2, "division": 3, "space": 4}
    _LEVEL_LABEL    = {0: "부지", 1: "시설", 2: "영역", 3: "과", 4: "세부"}
    _LEVEL_FILL_HEX = {0: "2F4F4F", 1: "4A7C8C", 2: "B8D4DC", 3: "E8F4F8", 4: None}
    _LEVEL_WHITE    = {0, 1}   # dark fill → white text

    area_rows = a["area_rows"]
    if area_rows:
        has_dept   = any(r_row.get("dept") for r_row in area_rows if isinstance(r_row, dict))
        n_cols_s5  = 7 if has_dept else 6
        s5_headers = ["레벨", "공간명", "기준면적(㎡)", "계획면적(㎡)", "비율(%)", "참고사항"]
        if has_dept:
            s5_headers.append("소관부서")

        ws5  = wb.create_sheet("5.면적표상세")
        row5 = _write_section_title(ws5, "5. 면적표 상세 프로그램", 1, span=n_cols_s5)
        row5 += 1
        ws5.freeze_panes  = "A3"
        ws5.print_title_rows = "1:2"

        row5 = _write_header(ws5, s5_headers, row5)

        for ar in area_rows:
            if not isinstance(ar, dict):
                continue
            rt      = ar.get("row_type") or "space"
            level   = _ROW_TYPE_LEVEL.get(rt, 4)
            lbl     = _LEVEL_LABEL.get(level, "")
            fill_hex = _LEVEL_FILL_HEX.get(level)
            lvl_fill = PatternFill("solid", fgColor=fill_hex) if fill_hex else None
            is_white = level in _LEVEL_WHITE
            is_sub   = bool(ar.get("is_subtotal"))

            indent       = level * 2
            name         = ar.get("name") or ""
            area_val_raw = ar.get("area") if ar.get("area") is not None else ar.get("subtotal_area")
            area_val     = _cell_safe(area_val_raw)
            ratio        = ar.get("ratio_pct")

            def _lv_font(**kw):
                """빌더: 레벨별 기본 폰트."""
                if is_white:
                    return Font(bold=True, color="FFFFFF", **kw)
                if is_sub:
                    return Font(bold=True, italic=True, **kw)
                return Font(**kw) if kw else None

            def _apply(cell, value=None, alignment=None, num_fmt=None):
                if value is not None:
                    cell.value = value
                if lvl_fill:
                    cell.fill = lvl_fill
                f = _lv_font()
                if f:
                    cell.font = f
                if alignment:
                    cell.alignment = alignment
                elif is_white:
                    cell.alignment = _center
                if num_fmt:
                    cell.number_format = num_fmt

            # Col A: 레벨명
            _apply(ws5.cell(row=row5, column=1), value=lbl, alignment=_center)

            # Col B: 공간명 (들여쓰기)
            name_display = ("　" * indent) + name  # full-width space for indent
            _apply(ws5.cell(row=row5, column=2), value=name_display,
                   alignment=Alignment(vertical="top", wrap_text=True))

            # Col C: 기준면적
            _apply(ws5.cell(row=row5, column=3), value=area_val,
                   alignment=_center if is_white else _num_right,
                   num_fmt=_NUM_FMT if isinstance(area_val_raw, (int, float)) else None)

            # Col D: 계획면적 (빈칸 — 설계자 입력용)
            _apply(ws5.cell(row=row5, column=4), alignment=_num_right, num_fmt=_NUM_FMT)

            # Col E: 비율
            ratio_cell = ws5.cell(row=row5, column=5)
            _apply(ratio_cell, value=_cell_safe(ratio), alignment=_num_right)
            if isinstance(ratio, (int, float)):
                ratio_cell.number_format = "0.0"

            # Col F: 참고사항
            _apply(ws5.cell(row=row5, column=6),
                   value=ar.get("notes") or "",
                   alignment=_wrap_top)

            # Col G: 소관부서 (optional)
            if has_dept:
                _apply(ws5.cell(row=row5, column=7),
                       value=ar.get("dept") or "",
                       alignment=_wrap_top)

            row5 += 1

        ws5.column_dimensions["A"].width = 8
        ws5.column_dimensions["B"].width = 40
        ws5.column_dimensions["C"].width = 15
        ws5.column_dimensions["D"].width = 15
        ws5.column_dimensions["E"].width = 10
        ws5.column_dimensions["F"].width = 30
        if has_dept:
            ws5.column_dimensions["G"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
