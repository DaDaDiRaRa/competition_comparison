"""
brief pipeline unit tests — validator + exporter (no LLM).
Run: python -m pytest tests/test_brief_pipeline.py -v
"""
import sys, os, tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import pytest
from services.brief_validator import validate_brief
from services.brief_checklist_exporter import to_markdown, to_xlsx, to_html

BRIEF_DATA = {
    '_quantitative': {
        'total_floor_area_sqm': 15000,
        'site_area_sqm': 5000,
        'building_coverage_ratio_pct': 60,
        'floor_area_ratio_pct': 300,
        'floors_above': 5,
        'parking_count': 100,
    },
    'brief_evaluation': {
        'evaluation_categories': [
            {'name': '설계 개념', 'points': 30, 'description': '창의성'},
            {'name': '기능성',    'points': 25, 'description': '공간구성'},
            {'name': '경관',     'points': 25, 'description': '도시환경'},
            {'name': '경제성',   'points': None, 'description': '유지관리'},   # null → medium flag
        ],
        'total_points': 100,
    },
    'brief_program': {
        'total_required_floor_area_sqm': 15000,
        'rooms': [
            {'name': '주민센터',   'required_area_sqm': 800, 'required_count': 1},
            {'name': '다목적강당', 'required_area_sqm': 500, 'required_count': 1},
        ],
    },
    'brief_design_guide': {
        'design_requirements': [
            '공공성과 개방성을 최우선으로 한다.',
            '친환경 설계 방식을 적용한다.',
        ],
    },
    'page_map': [
        {'page': 1, 'primary_type': 'BRIEF_OVERVIEW',    'confidence': 0.92},
        {'page': 2, 'primary_type': 'BRIEF_PROGRAM',     'confidence': 0.45},    # low → flag
        {'page': 3, 'primary_type': 'BRIEF_EVALUATION',  'confidence': 0.88},
        {'page': 4, 'primary_type': 'BRIEF_DESIGN_GUIDE','confidence': 0.80},
        {'page': 5, 'primary_type': 'BRIEF_SUBMISSION',  'confidence': 0.75},
    ],
}

REQUIREMENTS = {
    'requirements': [
        {'axis': 'concept_clarity',  'description': '설계 개념의 명확성', 'weight_pct': 30},
        {'axis': 'site_response',    'description': '대지와의 관계성',    'weight_pct': 25},
        {'axis': 'program_planning', 'description': '프로그램 적절성',    'weight_pct': 25},
    ],
    'special_requirements': ['장애인 편의시설 법정 기준 충족'],
}


class TestValidateBrief:
    def test_returns_validation_key(self):
        r = validate_brief(BRIEF_DATA, REQUIREMENTS)
        assert 'validation' in r

    def test_summary_keys(self):
        v = validate_brief(BRIEF_DATA, REQUIREMENTS)['validation']
        assert set(v['summary'].keys()) == {'high', 'medium', 'low'}

    def test_flags_are_list(self):
        v = validate_brief(BRIEF_DATA, REQUIREMENTS)['validation']
        assert isinstance(v['flags'], list)

    def test_flag_fields(self):
        v = validate_brief(BRIEF_DATA, REQUIREMENTS)['validation']
        for f in v['flags']:
            assert 'type' in f, f"missing 'type': {f}"
            assert 'severity' in f, f"missing 'severity': {f}"
            assert 'message' in f, f"missing 'message': {f}"
            assert 'location' in f, f"missing 'location': {f}"
            assert f['severity'] in ('high', 'medium', 'low')

    def test_null_points_detected(self):
        v = validate_brief(BRIEF_DATA, REQUIREMENTS)['validation']
        types = [f['type'] for f in v['flags']]
        assert 'points_mismatch' in types, "배점 null 항목 미감지"

    def test_low_confidence_detected(self):
        v = validate_brief(BRIEF_DATA, REQUIREMENTS)['validation']
        types = [f['type'] for f in v['flags']]
        assert 'low_confidence' in types, "저신뢰 페이지 미감지"

    def test_summary_counts_match_flags(self):
        v = validate_brief(BRIEF_DATA, REQUIREMENTS)['validation']
        flags = v['flags']
        summary = v['summary']
        assert summary['high']   == sum(1 for f in flags if f['severity'] == 'high')
        assert summary['medium'] == sum(1 for f in flags if f['severity'] == 'medium')
        assert summary['low']    == sum(1 for f in flags if f['severity'] == 'low')

    def test_checked_rules_present(self):
        v = validate_brief(BRIEF_DATA, REQUIREMENTS)['validation']
        assert 'checked_rules' in v
        assert len(v['checked_rules']) >= 4


class TestToMarkdown:
    def setup_method(self):
        v_result = validate_brief(BRIEF_DATA, REQUIREMENTS)
        self.bd = {**BRIEF_DATA, **v_result}
        self.validation = v_result['validation']

    def test_returns_string(self):
        md = to_markdown(self.bd, self.validation)
        assert isinstance(md, str)
        assert len(md) > 100

    def test_five_sections(self):
        md = to_markdown(self.bd, self.validation)
        assert '## 1. 사업 개요' in md
        assert '## 2. 면적 프로그램' in md
        assert '## 3. 심사기준' in md
        assert '## 4. 요구사항·설계 지침' in md
        assert '## 5. 검증 경고' in md

    def test_area_values_present(self):
        md = to_markdown(self.bd, self.validation)
        assert '15,000' in md or '15000' in md  # 연면적

    def test_evaluation_table(self):
        md = to_markdown(self.bd, self.validation)
        assert '설계 개념' in md
        assert '30' in md

    def test_flag_summary_line(self):
        md = to_markdown(self.bd, self.validation)
        # 검증 경고 섹션의 건수 요약 줄 (예: "높음: 1건 / 보통: 2건 / 낮음: 0건")
        assert '높음:' in md and '보통:' in md and '낮음:' in md

    def test_room_program(self):
        md = to_markdown(self.bd, self.validation)
        assert '주민센터' in md
        assert '다목적강당' in md


class TestToXlsx:
    def setup_method(self):
        v_result = validate_brief(BRIEF_DATA, REQUIREMENTS)
        self.bd = {**BRIEF_DATA, **v_result}
        self.validation = v_result['validation']

    def test_returns_bytes(self):
        b = to_xlsx(self.bd, self.validation)
        assert isinstance(b, bytes)
        assert len(b) > 1000

    def test_valid_xlsx(self):
        import openpyxl, io
        b = to_xlsx(self.bd, self.validation)
        wb = openpyxl.load_workbook(io.BytesIO(b))
        assert len(wb.sheetnames) == 4

    def test_sheet_names(self):
        import openpyxl, io
        b = to_xlsx(self.bd, self.validation)
        wb = openpyxl.load_workbook(io.BytesIO(b))
        names = wb.sheetnames
        assert '1.면적·프로그램' in names
        assert '2.심사기준'      in names
        assert '3.요구사항'      in names
        assert '4.검증경고'      in names

    def test_sheet1_has_data(self):
        import openpyxl, io
        b = to_xlsx(self.bd, self.validation)
        wb = openpyxl.load_workbook(io.BytesIO(b))
        ws = wb['1.면적·프로그램']
        values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert any('주민센터' in str(v) for v in values)

    def test_sheet2_score_present(self):
        import openpyxl, io
        b = to_xlsx(self.bd, self.validation)
        wb = openpyxl.load_workbook(io.BytesIO(b))
        ws = wb['2.심사기준']
        values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert any(v == 30 or v == '30' for v in values), "배점 30 미발견"

    def test_sheet4_severity_colors(self):
        import openpyxl, io
        b = to_xlsx(self.bd, self.validation)
        wb = openpyxl.load_workbook(io.BytesIO(b))
        ws = wb['4.검증경고']
        # 첫 번째 데이터 행에 fill이 있는지 확인
        for row in ws.iter_rows(min_row=4):
            fills = [c.fill.fgColor.rgb for c in row if c.fill and c.fill.fgColor]
            if any(f not in ('00000000', '000000') for f in fills):
                break
        # 경고가 있으면 색 fill이 적용되어야 함
        flag_count = sum(self.validation['summary'].values())
        if flag_count > 0:
            # 경고 행이 존재하면 최소 1개 이상의 비기본 색이 있어야 함
            all_fills = set()
            for row in ws.iter_rows(min_row=5):
                for c in row:
                    if c.fill and c.fill.fgColor:
                        all_fills.add(c.fill.fgColor.rgb)
            non_default = all_fills - {'00000000', '000000', 'FFFFFFFF', None}
            assert len(non_default) > 0, f"severity 색상 미적용: fills={all_fills}"


class TestToHtml:
    def setup_method(self):
        v_result = validate_brief(BRIEF_DATA, REQUIREMENTS)
        self.bd = {**BRIEF_DATA, **v_result}
        self.validation = v_result['validation']

    def test_returns_wellformed_document(self):
        h = to_html(self.bd, self.validation)
        assert isinstance(h, str)
        assert h.startswith('<!DOCTYPE html>')
        assert h.rstrip().endswith('</html>')
        # 태그 균형 (void 제외)
        from html.parser import HTMLParser
        class _B(HTMLParser):
            VOID = {'meta', 'br', 'hr', 'img', 'input', 'link', 'col'}
            def __init__(self): super().__init__(); self.stk = []; self.bad = 0
            def handle_starttag(self, t, a):
                if t not in self.VOID: self.stk.append(t)
            def handle_endtag(self, t):
                if t in self.VOID: return
                if self.stk and self.stk[-1] == t: self.stk.pop()
                else: self.bad += 1
        b = _B(); b.feed(h)
        assert b.stk == [] and b.bad == 0

    def test_five_sections(self):
        h = to_html(self.bd, self.validation)
        for title in ('사업 개요', '면적 프로그램', '심사기준', '요구사항·설계 지침', '검증 경고'):
            assert title in h

    def test_content_present(self):
        h = to_html(self.bd, self.validation)
        assert '주민센터' in h          # 면적 프로그램
        assert '30' in h                # 배점

    def test_escapes_html_in_data(self):
        bd = {**self.bd, '_brief_meta': {'brief_name': '<script>alert(1)</script>'}}
        h = to_html(bd, self.validation)
        assert '<script>alert(1)</script>' not in h
        assert '&lt;script&gt;' in h

    # ── AI 종합 해설 섹션 (insight 파라미터, Unit 4) ───────────────────────────
    _INSIGHT = {
        "synthesis_summary": "이 지침서는 기능·동선 해결에 무게를 둔다.",
        "key_emphases": [
            {"topic": "동선 분리", "signal_strength": "strong",
             "signals": ["배치계획 배점 1순위", "동선계획 14항목"],
             "basis": ["p.20", "배치계획"], "note": "감염·보안 동선 분리 반복 요구"},
        ],
        "must_not_miss": [{"item": "재직증명서 미제출 시 실격", "basis": "p.18"}],
        "hidden_constraints": [{"issue": "용적률은 심의로 결정", "basis": "p.5", "note": "법정 한계 아님"}],
        "reading_guide": ["배점이 정성평가 비중 큼"],
        "data_confidence": "high",
        "caveats": ["배점표 일부 추출 불완전"],
    }

    @staticmethod
    def _tag_balance(h: str) -> tuple[list, int]:
        from html.parser import HTMLParser

        class _B(HTMLParser):
            VOID = {'meta', 'br', 'hr', 'img', 'input', 'link', 'col'}

            def __init__(self):
                super().__init__(); self.stk = []; self.bad = 0

            def handle_starttag(self, t, a):
                if t not in self.VOID:
                    self.stk.append(t)

            def handle_endtag(self, t):
                if t in self.VOID:
                    return
                if self.stk and self.stk[-1] == t:
                    self.stk.pop()
                else:
                    self.bad += 1

        b = _B(); b.feed(h)
        return b.stk, b.bad

    def test_no_insight_section_by_default(self):
        h = to_html(self.bd, self.validation)
        assert '지침서 종합 해설' not in h
        assert 'id="insight"' not in h

    def test_insight_section_renders(self):
        h = to_html(self.bd, self.validation, insight=self._INSIGHT)
        assert 'id="insight"' in h
        assert '지침서 종합 해설' in h
        assert '이 지침서는 기능·동선 해결에 무게를 둔다.' in h   # synthesis
        assert '동선 분리' in h                                   # key_emphasis topic
        assert '재직증명서 미제출 시 실격' in h                    # must_not_miss
        assert 'p.20' in h                                       # basis 인용
        assert '당락 예측이 아닙니다' in h                         # disclaimer
        assert '#insight">해설</a>' in h                          # nav 항목

    def test_insight_section_wellformed(self):
        h = to_html(self.bd, self.validation, insight=self._INSIGHT)
        stk, bad = self._tag_balance(h)
        assert stk == [] and bad == 0

    def test_insight_escapes_html(self):
        evil = {**self._INSIGHT, "synthesis_summary": "<script>alert(1)</script>"}
        h = to_html(self.bd, self.validation, insight=evil)
        assert '<script>alert(1)</script>' not in h
        assert '&lt;script&gt;' in h

    def test_insight_minimal_is_wellformed(self):
        """배지만 있고 하위 블록이 비어도 섹션은 깨지지 않는다 (graceful)."""
        h = to_html(self.bd, self.validation, insight={"data_confidence": "low"})
        assert 'id="insight"' in h
        assert '근거 낮음' in h
        stk, bad = self._tag_balance(h)
        assert stk == [] and bad == 0


class TestSiteLawSection:
    """대지·법적 골격 섹션 (_site_context) → html/md 렌더. LLM 0. #7 — 제안서 없이도 대지·법 표시."""

    def _bd(self):
        law = [
            {"site_id": "부지1", "address": "영등포",
             "envelope": {"bcr_limit_pct": 60.0, "far_limit_pct": 400.0},
             "height_solar": {"north_setback_m": None, "shadow_applies": False,
                              "shadow_setback_rule": None, "shadow_min_setback_m": None,
                              "road_height_limit_m": None, "parcel_north_depth_m": None},
             "reviews_required": [{"name": "건축위원회 심의"}, {"name": "경관심의"}],
             "has_required_review": True, "low_confidence": False, "source_notes": {},
             "limit_mismatch": [{"field": "용적률", "brief_pct": 460, "diagnose_limit_pct": 400.0}]},
            {"site_id": "부지주거", "address": "하안주공",
             "envelope": {"bcr_limit_pct": 60.0, "far_limit_pct": 250.0},
             "height_solar": {"north_setback_m": None, "shadow_applies": True,
                              "shadow_setback_rule": "높이/2 후퇴", "shadow_min_setback_m": 65.0,
                              "road_height_limit_m": None, "parcel_north_depth_m": None},
             "reviews_required": [{"name": "도시계획위원회 심의"}], "has_required_review": True,
             "low_confidence": True, "source_notes": {}, "limit_mismatch": []},
        ]
        return {
            "_brief_meta": {"facility_type": "public", "brief_name": "테스트"},
            "_site_context": {
                "matched_address": "영등포",
                "analysis": {"overall_summary": "저층 주거 밀집 시가지",
                             "orientation": "남측 접도", "road_access": "당산로27길"},
                "measured": {"design_drivers": [{"name": "1인가구 대응"}, {"name": "방재"}]},
                "law_diagnosis": law,
            },
        }

    _V = {"flags": [], "summary": {}}

    def test_html_renders_section(self):
        h = to_html(self._bd(), self._V)
        assert 'id="sitelaw"' in h and "대지 · 법적 골격" in h
        assert "필수 심의" in h and "건축위원회 심의" in h
        assert "정북 일조" in h and "필요이격 65m" in h        # 주거지역 정북(패치 후 shadow_min)
        assert "brief 수치 재확인" in h                        # limit_mismatch 경고
        assert "남측 접도" in h and "실측 설계 드라이버" in h   # 대지 요약 + 터읽기

    def test_md_renders_section(self):
        m = to_markdown(self._bd(), self._V)
        assert "## 0.5 대지 · 법적 골격" in m
        assert "필수 심의" in m and "필요이격 65m" in m
        assert "용적률 재확인" in m

    def test_absent_when_no_site_context(self):
        bd = {"_brief_meta": {"facility_type": "public"}}
        assert 'id="sitelaw"' not in to_html(bd, self._V)
        assert "0.5 대지" not in to_markdown(bd, self._V)


class TestSiteLawXlsx:
    """대지·법적 골격 xlsx 시트 (#7 마무리). site_context 있으면 시트, 없으면 생략."""

    def test_xlsx_sheet_renders(self):
        import io as _io, openpyxl as _ox
        bd = TestSiteLawSection()._bd()
        wb = _ox.load_workbook(_io.BytesIO(to_xlsx(bd, {"flags": [], "summary": {}})))
        assert "대지·법적 골격" in wb.sheetnames
        ws = wb["대지·법적 골격"]
        blob = " | ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        assert "건축위원회 심의" in blob and "도시계획위원회 심의" in blob
        assert "필요이격 65m" in blob        # 주거 정북
        assert "남측 접도" in blob            # 대지 요약
        assert "용적률" in blob and "460" in blob   # mismatch 재확인

    def test_xlsx_sheet_absent_without_site_context(self):
        import io as _io, openpyxl as _ox
        wb = _ox.load_workbook(_io.BytesIO(to_xlsx({"_brief_meta": {"facility_type": "public"}},
                                                   {"flags": [], "summary": {}})))
        assert "대지·법적 골격" not in wb.sheetnames


# ── OMA식 면적 프로그램 비례 스택 다이어그램 (LLM 0 · 결정론 SVG) ──────────────────
class TestProgramAreaStack:
    from services.brief_checklist_exporter import (
        _program_stack_blocks as _blocks, _program_area_stack_svg as _svg,
        _to_area_float as _tof,
    )

    def test_facility_subtotal_no_double_count(self):
        # 시설 subtotal 이 있으면 하위 space 를 더하지 않는다 (이중집계 방지).
        a = {'area_rows': [
            {'row_type': 'site_total', 'name': '전체', 'subtotal_area': 300000},
            {'row_type': 'facility', 'name': '서고', 'subtotal_area': 200000},
            {'row_type': 'space', 'name': '서고1', 'area': 150000},
            {'row_type': 'facility', 'name': '주차', 'subtotal_area': 100000},
        ], 'area_table': [], 'rooms': [], 'zones': []}
        blocks, denom, note = TestProgramAreaStack._blocks(a)
        assert blocks == [('서고', 200000.0), ('주차', 100000.0)]
        assert denom == 300000.0          # site_total 은 블록 아님, denom=블록합
        assert note == ''

    def test_facility_children_summed_when_no_subtotal(self):
        a = {'area_rows': [
            {'row_type': 'facility', 'name': 'A', 'subtotal_area': None},
            {'row_type': 'space', 'name': 'a1', 'area': 30},
            {'row_type': 'space', 'name': 'a2', 'area': 20},
            {'row_type': 'facility', 'name': 'B', 'subtotal_area': 40},
        ], 'area_table': [], 'rooms': [], 'zones': []}
        blocks, denom, _ = TestProgramAreaStack._blocks(a)
        assert blocks == [('A', 50.0), ('B', 40.0)]

    def test_gating_under_two_blocks(self):
        a = {'area_rows': [{'row_type': 'facility', 'name': 'x', 'subtotal_area': None}],
             'area_table': [], 'rooms': [], 'zones': []}
        assert TestProgramAreaStack._blocks(a) == ([], None, '')

    def test_area_table_fallback_and_string_coerce(self):
        a = {'area_rows': [], 'rooms': [], 'zones': [], 'area_table': [
            {'group_name': 'A', 'total_area_sqm': '12,000 ㎡'},
            {'group_name': 'B', 'items': [{'area_sqm': 3000}, {'area_sqm': 2000}]},
        ]}
        blocks, denom, _ = TestProgramAreaStack._blocks(a)
        assert blocks == [('A', 12000.0), ('B', 5000.0)]
        assert denom == 17000.0

    def test_many_blocks_collapse_to_other(self):
        rows = [{'row_type': 'facility', 'name': f'F{i}', 'subtotal_area': 100 - i}
                for i in range(20)]
        a = {'area_rows': rows, 'area_table': [], 'rooms': [], 'zones': []}
        blocks, _, note = TestProgramAreaStack._blocks(a)
        assert len(blocks) == 14                      # 상위 13 + 기타
        assert blocks[-1][0].startswith('기타')
        assert '기타' in note

    def test_to_area_float(self):
        tof = TestProgramAreaStack._tof
        assert tof('12,345.6 ㎡') == 12345.6
        assert tof(5000) == 5000.0
        assert tof(None) is None and tof('-5') is None and tof('abc') is None and tof(0) is None

    def test_svg_escapes_and_side_tab(self):
        blocks = [('큰방', 100000.0), ('행정<b>x</b>', 34000.0), ('열람실', 300.0)]
        svg = TestProgramAreaStack._svg(blocks, sum(v for _, v in blocks), '연면적 X ㎡', '')
        assert '<svg' in svg and 'aria-label="면적 프로그램 비례 다이어그램"' in svg
        assert '행정&lt;b&gt;x&lt;/b&gt;' in svg     # 이름 escape
        assert '<b>x</b>' not in svg
        assert '열람실' in svg and '<line' in svg     # 얇은 블록 → 옆 탭 리더선

    def test_diagram_appears_in_to_html(self):
        # BRIEF_DATA 는 rooms 2개 → 다이어그램 렌더.
        v = validate_brief(BRIEF_DATA, REQUIREMENTS)
        h = to_html({**BRIEF_DATA, **v}, v['validation'])
        assert 'aria-label="면적 프로그램 비례 다이어그램"' in h

    def test_multisite_scoping_excludes_resummary_and_subtotals(self):
        # 다부지: '부지' site_total 아래 요약 시설만 채택. ①②③ 재집계 헤더/시설·소계 행 배제,
        # 부지 간 동일 시설(주차)은 합산 (영등포 통합신청사 교훈).
        a = {'area_rows': [
            {'row_type': 'site_total', 'name': '총 합계 (지상+지하)', 'subtotal_area': 600},
            {'row_type': 'site_total', 'name': '부지1(당산 385) 합계 계', 'subtotal_area': 400},
            {'row_type': 'facility', 'name': '구청', 'subtotal_area': 300},
            {'row_type': 'facility', 'name': '주차(지하)', 'subtotal_area': 100},
            {'row_type': 'site_total', 'name': '부지2(당산 370) 합계 계', 'subtotal_area': 200},
            {'row_type': 'facility', 'name': '보건소', 'subtotal_area': 150},
            {'row_type': 'facility', 'name': '주차(지하)', 'subtotal_area': 50},
            {'row_type': 'site_total', 'name': '① 계(지상+지하)', 'subtotal_area': 400},  # 재집계 헤더
            {'row_type': 'facility', 'name': '① 구청', 'subtotal_area': 300},              # 재집계 시설
            {'row_type': 'space', 'name': '청장실', 'area': 99},
            {'row_type': 'facility', 'name': '구청 전용 계', 'subtotal_area': 300},         # 소계
        ], 'area_table': [], 'rooms': [], 'zones': []}
        blocks, denom, _ = TestProgramAreaStack._blocks(a)
        assert blocks == [('구청', 300.0), ('주차(지하)', 150.0), ('보건소', 150.0)]
        assert denom == 600.0                          # 연면적 총합과 일치, 재집계 미포함

    def test_clean_strips_decoration_merges_and_drops_subtotals(self):
        from services.brief_checklist_exporter import _clean_program_blocks as _clean
        out = _clean([
            ('▣ 공용', 100), ('공용', 50),             # 장식 제거 → 동일 라벨 합산
            ('소계', 999), ('보건소 전용 계', 999), ('총 공용+설비', 999),  # 소계·재집계 제외
            ('A', 0), ('B', -5),                         # 비양수 제외
            ('C', 30),
        ])
        assert out == [('공용', 150.0), ('C', 30.0)]

    def test_is_subtotal_and_norm_name(self):
        from services.brief_checklist_exporter import _is_subtotal_name as sub, _norm_prog_name as nm
        assert sub('소계') and sub('구청 전용 계') and sub('총 공용') and sub('합계')
        assert not sub('구청') and not sub('부속 주차장(지하)') and not sub('공공커뮤니티지원센터')
        assert nm('▣ 공용') == '공용' and nm('■ 구청') == '구청' and nm('구청') == '구청'
