"""xlsx 생성 결과 내용 확인 스크립트."""
import sys, io, os
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from services.brief_validator import validate_brief
from services.brief_checklist_exporter import to_markdown, to_xlsx
import openpyxl

bd = {
    '_quantitative': {
        'total_floor_area_sqm': 15000, 'site_area_sqm': 5000,
        'building_coverage_ratio_pct': 60, 'floor_area_ratio_pct': 300,
        'floors_above': 5, 'parking_count': 100,
    },
    'brief_evaluation': {
        'evaluation_categories': [
            {'name': '설계 개념', 'points': 30, 'description': '창의성'},
            {'name': '기능성',    'points': 25, 'description': '공간구성'},
            {'name': '경관',      'points': 25, 'description': '도시환경'},
            {'name': '경제성',    'points': None, 'description': '유지관리'},
        ],
        'total_points': 100,
        'disqualification_criteria': ['분량 초과 시 실격', '기명 작품 실격'],
    },
    'brief_program': {
        'total_required_floor_area_sqm': 15000,
        'rooms': [
            {'name': '주민센터',   'required_area_sqm': 800,  'required_count': 1, 'floor': '1F'},
            {'name': '다목적강당', 'required_area_sqm': 500,  'required_count': 1, 'floor': '2F'},
            {'name': '옥상정원',   'required_area_sqm': 200,  'required_count': 1, 'floor': 'RF'},
        ],
        'zones': [
            {'name': '공공존', 'area_sqm': 1300},
            {'name': '서비스존', 'area_sqm': 200},
        ],
    },
    'brief_design_guide': {
        'concept_direction': '주민과 함께하는 열린 공공공간',
        'design_requirements': ['공공성과 개방성을 최우선으로 한다.', '친환경 설계 방식을 적용한다.'],
        'prohibited_items': ['전면 유리 커튼월 사용 금지'],
        'sustainability_requirements': ['녹색건축 인증 취득 권장'],
    },
    'page_map': [
        {'page': 1, 'primary_type': 'BRIEF_OVERVIEW',     'confidence': 0.92},
        {'page': 2, 'primary_type': 'BRIEF_PROGRAM',      'confidence': 0.45},
        {'page': 3, 'primary_type': 'BRIEF_EVALUATION',   'confidence': 0.88},
        {'page': 4, 'primary_type': 'BRIEF_DESIGN_GUIDE', 'confidence': 0.80},
        {'page': 5, 'primary_type': 'BRIEF_SUBMISSION',   'confidence': 0.75},
    ],
}
reqs = {
    'requirements': [
        {'axis': 'concept_clarity',  'description': '설계 개념의 명확성', 'weight_pct': 30},
        {'axis': 'site_response',    'description': '대지와의 관계성',    'weight_pct': 25},
        {'axis': 'program_planning', 'description': '프로그램 적절성',    'weight_pct': 25},
    ],
    'special_requirements': ['장애인 편의시설 법정 기준 100% 충족', '주민 편의 동선 최우선 확보'],
}

v_result = validate_brief(bd, reqs)
bd.update(v_result)
validation = v_result['validation']

print('=== 검증 경고 ===')
for f in validation['flags']:
    sev = f['severity']
    print(f'  [{sev}] {f["type"]}: {f["message"][:70]}')
print(f'요약: {validation["summary"]}')
print()

xlsx_bytes = to_xlsx(bd, validation)
out = r'C:\Temp\test_brief_checklist.xlsx'
os.makedirs(r'C:\Temp', exist_ok=True)
with open(out, 'wb') as fh:
    fh.write(xlsx_bytes)
print(f'xlsx 저장: {out} ({len(xlsx_bytes):,} bytes)')
print()

wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
for name in wb.sheetnames:
    ws = wb[name]
    print(f'[{name}] {ws.max_row}행 x {ws.max_column}열')
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True):
        if any(v is not None for v in row):
            print(f'  {row}')
    print()

print('=== Markdown 첫 40줄 ===')
md = to_markdown(bd, validation)
for line in md.split('\n')[:40]:
    print(line)
